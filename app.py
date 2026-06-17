import streamlit as st
import anthropic
import pandas as pd
from pypdf import PdfReader
import io
import json
import html as _html
import plotly.express as px
import plotly.graph_objects as go
from collections import Counter
import streamlit_authenticator as stauth
from supabase import create_client
from datetime import datetime, timezone


def esc(val) -> str:
    """Escapa HTML en valores de respuestas de la API o input externo."""
    return _html.escape(str(val))


def safe_url(url) -> str:
    """Valida que una URL use http/https antes de usarla como href."""
    u = str(url).strip()
    return _html.escape(u) if u.startswith(("http://", "https://")) else "#"

st.set_page_config(page_title="Nexora", page_icon="N", layout="wide")

# â”€â”€ AUTENTICACION â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def _cargar_credentials() -> dict:
    usuarios = st.secrets.get("auth", {}).get("usuarios", {})
    if not usuarios:
        st.error("No se encontraron usuarios en secrets.toml → [auth.usuarios.<nombre>]")
        st.stop()
    return {
        "usernames": {
            nombre: {"name": data["name"], "password": data["password"]}
            for nombre, data in usuarios.items()
        }
    }

cookie_secret = st.secrets.get("COOKIE_SECRET", "")
if not cookie_secret:
    st.error("Falta COOKIE_SECRET en secrets.toml")
    st.stop()
authenticator = stauth.Authenticate(
    _cargar_credentials(), "nexora_cookie", cookie_secret, cookie_expiry_days=30
)

# â”€â”€ CSS SUNLIT SLATE (clara y calida) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap');

  /*
   * SKILLS APPLIED (11 total):
   * tasteskill · emilkowalski · ihlamury/linear · ihlamury/stripe
   * ihlamury/vercel · ihlamury/notion · anthropics/frontend-design
   * nextlevelbuilder/ui-ux-pro-max · delphi-ai/animate
   * kylezantos/design-motion-principles · vercel-labs/web-design-guidelines
   */

  :root {
    --nx-sans: 'Outfit', sans-serif;
    --nx-mono: 'IBM Plex Mono', monospace;

    /* Stripe bg, Vercel surfaces */
    --bg:      #F2F6F9;
    --surface: #FFFFFF;
    --sidebar: #1A2736;
    --accent:  #3574A8;
    --accent-lt: #EBF4FB;

    /* Text hierarchy — Vercel 3-level */
    --text-primary:   #0D1B2A;
    --text-secondary: #4A5568;
    --text-muted:     #8896A4;

    /* Borders — Notion 1px subtlety */
    --border:    #E2E8F0;
    --border-md: #CBD5E1;

    /* animate-skill: 3 easing curves for 3 motion types */
    --ease-out:        cubic-bezier(0.23, 1, 0.32, 1);    /* entrances */
    --ease-out-cubic:  cubic-bezier(0.33, 1, 0.68, 1);    /* snappier entrances */
    --ease-in-out:     cubic-bezier(0.645, 0.045, 0.355, 1); /* elements moving on screen */
    --ease-in:         cubic-bezier(0.55, 0, 1, 0.45);    /* exits */

    /* Focus ring — Linear/web-design-guidelines: 2px, accent-color */
    --focus-ring: 0 0 0 2px rgba(53,116,168,0.2);

    /* Linear: 4px grid, strict 6/8/12 radius scale */
    --r-sm: 6px;
    --r-md: 8px;
    --r-lg: 12px;

    /* Slate 3D multi-layer shadows */
    --sh-card: 0 0 0 0.5px rgba(13,27,42,0.04), 0 1px 3px rgba(0,0,0,0.04), 0 6px 20px rgba(13,27,42,0.06);
    --sh-card-up: 0 0 0 1px rgba(13,27,42,0.06), 0 4px 14px rgba(13,27,42,0.08), 0 18px 40px rgba(13,27,42,0.1);
    --sh-btn: 0 1px 2px rgba(13,27,42,0.14), 0 4px 12px rgba(13,27,42,0.12), 0 0 0 0.5px rgba(13,27,42,0.06);
    --sh-btn-up: 0 2px 6px rgba(13,27,42,0.18), 0 10px 28px rgba(13,27,42,0.14), 0 0 0 0.5px rgba(13,27,42,0.06);
  }

  html,body,[class*="css"] { font-family: var(--nx-sans) !important; }

  /* ── APP BACKGROUND ── */
  .stApp,.stApp>div,[data-testid="stAppViewContainer"],
  [data-testid="stAppViewContainer"]>section,
  [data-testid="stAppViewContainer"]>section>div,
  .main,.block-container { background-color: var(--bg) !important; overflow-x: hidden !important; }
  [data-testid="stVerticalBlock"],[data-testid="stHorizontalBlock"] { background: transparent !important; }
  .stApp,.stApp p,.stApp span,.stApp div,.stApp label,
  .stMarkdown,.stMarkdown p { color: var(--text-primary) !important; }

  /* web-design-guidelines: tap highlight for mobile */
  button,a,[role="button"],.stButton>button,.stDownloadButton>button {
    -webkit-tap-highlight-color: transparent !important;
    touch-action: manipulation !important; }

  /* ── SIDEBAR — todos los niveles de div anidados ── */
  [data-testid="stSidebar"],
  [data-testid="stSidebar"]>div,
  [data-testid="stSidebar"]>div>div,
  [data-testid="stSidebar"]>div>div>div,
  section[data-testid="stSidebar"],
  [data-testid="stSidebarContent"],
  [data-testid="stSidebarUserContent"] {
    background-color: var(--sidebar) !important; }
  [data-testid="stSidebar"] {
    border-right: 1px solid rgba(255,255,255,0.05) !important;
    box-shadow: 8px 0 40px rgba(0,0,0,0.28), inset -1px 0 0 rgba(255,255,255,0.04) !important; }
  /* Padding interior del sidebar — evita q widgets queden pegados al borde */
  [data-testid="stSidebarUserContent"],
  [data-testid="stSidebarUserContent"] > div,
  [data-testid="stSidebarContent"] > div > div:last-child {
    padding-left: 10px !important;
    padding-right: 10px !important; }
  /* Sidebar: texto claro — incluye div y strong para neutralizar .stApp div !important */
  [data-testid="stSidebar"] label,
  [data-testid="stSidebar"] p,
  [data-testid="stSidebar"] div,
  [data-testid="stSidebar"] span,
  [data-testid="stSidebar"] strong { color: rgba(240,246,252,0.72) !important; }
  [data-testid="stSidebar"] label { font-weight: 400 !important; }
  [data-testid="stSidebar"] h1,[data-testid="stSidebar"] h2,
  [data-testid="stSidebar"] h3 { color: #F0F6FC !important; }
  /* Brand hierarchy: clases nx-sb-* tienen mayor especificidad → vencen al div generico */
  [data-testid="stSidebar"] .nx-sb-primary { color: #F0F6FC !important; font-weight: 700 !important; }
  [data-testid="stSidebar"] .nx-sb-sub     { color: rgba(240,246,252,0.48) !important; }
  [data-testid="stSidebar"] .nx-sb-muted   { color: rgba(240,246,252,0.48) !important; }
  [data-testid="stSidebar"] .nx-sb-name    { color: rgba(240,246,252,0.90) !important; font-weight: 600 !important; }
  /* Dividers — ultra-sutiles, solo 1px con 6% blanco */
  [data-testid="stSidebar"] hr,
  [data-testid="stSidebar"] [data-testid="stDivider"] {
    background-color: rgba(255,255,255,0.06) !important;
    border: none !important; border-top: none !important;
    height: 1px !important; margin: 4px 0 !important; }
  [data-testid="stSidebar"] [data-testid="stDivider"] > * {
    background-color: rgba(255,255,255,0.06) !important;
    border: none !important; }

  /* ── NAV — oculta el label "Módulo activo" del radio ── */
  [data-testid="stSidebar"] .stRadio > label {
    display: none !important; }
  /* ── NAV — motion-principles: nav is high-freq → keep transitions minimal ── */
  .stRadio [role="radiogroup"] { gap: 1px !important; }
  .stRadio label {
    background: transparent !important; border: none !important;
    border-left: 2px solid transparent !important;
    border-radius: 0 var(--r-sm) var(--r-sm) 0 !important;
    padding: 7px 12px !important; margin: 0 !important;
    transition: background 120ms var(--ease-out), border-color 120ms var(--ease-out) !important; }
  @media (hover: hover) and (pointer: fine) {
    .stRadio label:hover { background: rgba(255,255,255,0.07) !important; } }
  .stRadio label:has(input:checked) {
    background: rgba(53,116,168,0.22) !important; border-left-color: #5B9EC9 !important;
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.06) !important; }
  .stRadio label>div:first-child { display: none !important; }
  .stRadio label p {
    font-size: 12px !important; color: rgba(240,246,252,0.88) !important;
    font-weight: 400 !important; white-space: nowrap !important;
    overflow: hidden !important; text-overflow: ellipsis !important;
    transition: color 120ms var(--ease-out) !important; }
  @media (hover: hover) and (pointer: fine) {
    .stRadio label:hover p { color: #F0F6FC !important; } }
  .stRadio label:has(input:checked) p { color: #F0F6FC !important; font-weight: 500 !important; }

  /* ── HISTORIAL ── */
  .nx-hist-item { display: flex; align-items: center; gap: 8px; padding: 5px 8px;
    border-radius: var(--r-sm); border-left: 2px solid transparent; cursor: pointer;
    transition: background 120ms var(--ease-out); margin-bottom: 1px; }
  @media (hover: hover) and (pointer: fine) {
    .nx-hist-item:hover { background: rgba(255,255,255,0.07); } }
  .nx-hist-item.activo { background: rgba(53,116,168,0.16); border-left-color: #5B9EC9; }
  .nx-hist-titulo { flex: 1; min-width: 0; font-size: 11.5px; color: rgba(240,246,252,0.65);
    line-height: 1.3; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
  .nx-hist-item.activo .nx-hist-titulo { color: #F0F6FC; font-weight: 500; }
  .nx-hist-fecha { font-family: var(--nx-mono); font-size: 9px; color: rgba(240,246,252,0.28);
    letter-spacing: 0.5px; flex-shrink: 0; white-space: nowrap; }
  .nx-hist-empty { font-size: 11.5px; color: rgba(240,246,252,0.55); padding: 8px 8px; font-style: italic; }
  div[data-testid="stSidebar"] .nx-hist-btn button {
    border: none !important; background: transparent !important; padding: 2px 4px !important;
    min-height: 0 !important; height: auto !important; color: rgba(240,246,252,0.28) !important;
    font-size: 10px !important; letter-spacing: 0 !important; text-transform: none !important; }
  div[data-testid="stSidebar"] .nx-hist-btn button:hover { color: #5B9EC9 !important; }

  /* ── INPUTS — web-design-guidelines: no outline-none without replacement ── */
  input[type="text"],input[type="password"],.stTextInput input,.stTextArea textarea {
    background: var(--surface) !important; color: var(--text-primary) !important;
    border: 1px solid var(--border) !important; border-radius: var(--r-md) !important;
    font-size: 13.5px !important; font-family: var(--nx-sans) !important;
    box-shadow: 0 1px 2px rgba(0,0,0,0.04) !important;
    transition: border-color 160ms var(--ease-out), box-shadow 160ms var(--ease-out) !important; }
  input:focus-visible,textarea:focus-visible {
    border-color: var(--accent) !important;
    box-shadow: var(--focus-ring), 0 1px 2px rgba(0,0,0,0.04) !important;
    outline: none !important; }
  /* web-design-guidelines: never block paste */
  input,textarea { -webkit-user-select: text !important; user-select: text !important; }
  input::placeholder,textarea::placeholder { color: #94A3B8 !important; }
  .stTextArea textarea { border-radius: var(--r-md) !important; }
  .stSelectbox>div>div,.stSelectbox [data-baseweb="select"]>div {
    background: var(--surface) !important; color: var(--text-primary) !important;
    border: 1px solid var(--border) !important; border-radius: var(--r-md) !important;
    box-shadow: 0 1px 2px rgba(0,0,0,0.04) !important; }
  .stSelectbox [data-baseweb="select"] span,
  .stSelectbox [data-baseweb="select"] div { color: var(--text-primary) !important; }
  [data-baseweb="popover"] ul,[data-baseweb="popover"] li,[data-baseweb="menu"] {
    background: var(--surface) !important; color: var(--text-primary) !important;
    border-radius: var(--r-md) !important; box-shadow: var(--sh-card-up) !important; }
  [data-baseweb="option"]:hover { background: var(--accent-lt) !important; }
  .stMultiSelect>div>div {
    background: var(--surface) !important; border: 1px solid var(--border) !important;
    border-radius: var(--r-md) !important; color: var(--text-primary) !important; }

  /* ── BUTTONS — animate-skill: press scale, named transitions ── */
  .stButton>button[kind="primary"] {
    background: #0D1B2A !important; border: none !important;
    border-radius: var(--r-md) !important; color: #F0F6FC !important;
    font-size: 13px !important; font-weight: 500 !important;
    font-family: var(--nx-sans) !important; letter-spacing: 0.15px !important;
    text-transform: none !important; padding: 9px 20px !important;
    box-shadow: var(--sh-btn) !important;
    transition: background 160ms var(--ease-out),
                transform 160ms var(--ease-out),
                box-shadow 160ms var(--ease-out) !important; }
  @media (hover: hover) and (pointer: fine) {
    .stButton>button[kind="primary"]:hover {
      background: #1A2736 !important; transform: translateY(-2px) !important;
      box-shadow: var(--sh-btn-up) !important; } }
  .stButton>button[kind="primary"]:active {
    transform: scale(0.97) translateY(0) !important;
    box-shadow: 0 1px 2px rgba(13,27,42,0.12) !important; }
  .stButton>button[kind="primary"]:focus-visible {
    outline: 2px solid var(--accent) !important; outline-offset: 2px !important; }
  .stButton>button[kind="primary"]:disabled {
    opacity: 0.5 !important; cursor: not-allowed !important; transform: none !important; }

  .stButton>button {
    border-radius: var(--r-md) !important; border: 1px solid var(--border-md) !important;
    color: #334155 !important; background: var(--surface) !important;
    font-size: 13px !important; font-weight: 500 !important;
    font-family: var(--nx-sans) !important; letter-spacing: 0.1px !important;
    text-transform: none !important; padding: 8px 18px !important;
    box-shadow: 0 1px 2px rgba(0,0,0,0.06), 0 0 0 0.5px rgba(13,27,42,0.04) !important;
    transition: border-color 160ms var(--ease-out),
                color 160ms var(--ease-out),
                background 160ms var(--ease-out),
                transform 160ms var(--ease-out),
                box-shadow 160ms var(--ease-out) !important; }
  @media (hover: hover) and (pointer: fine) {
    .stButton>button:hover {
      border-color: var(--accent) !important; color: var(--accent) !important;
      background: var(--accent-lt) !important; transform: translateY(-1px) !important;
      box-shadow: 0 1px 3px rgba(53,116,168,0.1), 0 4px 14px rgba(53,116,168,0.1),
        0 0 0 0.5px rgba(53,116,168,0.12) !important; } }
  .stButton>button:active { transform: scale(0.97) !important; box-shadow: none !important; }
  .stButton>button:focus-visible {
    outline: 2px solid var(--accent) !important; outline-offset: 2px !important; }
  .stButton>button:disabled {
    opacity: 0.5 !important; cursor: not-allowed !important; transform: none !important; }

  .stDownloadButton>button {
    background: var(--surface) !important; border: 1.5px solid var(--accent) !important;
    border-radius: var(--r-md) !important; color: var(--accent) !important;
    font-size: 13px !important; font-weight: 500 !important;
    font-family: var(--nx-sans) !important; text-transform: none !important;
    box-shadow: 0 1px 3px rgba(53,116,168,0.1), 0 0 0 0.5px rgba(53,116,168,0.08) !important;
    transition: background 160ms var(--ease-out),
                color 160ms var(--ease-out),
                transform 160ms var(--ease-out),
                box-shadow 160ms var(--ease-out) !important; }
  @media (hover: hover) and (pointer: fine) {
    .stDownloadButton>button:hover {
      background: var(--accent) !important; color: var(--surface) !important;
      transform: translateY(-2px) !important;
      box-shadow: 0 4px 14px rgba(53,116,168,0.3), 0 10px 28px rgba(53,116,168,0.2) !important; } }
  .stDownloadButton>button:active { transform: scale(0.97) !important; }
  .stDownloadButton>button:focus-visible {
    outline: 2px solid var(--accent) !important; outline-offset: 2px !important; }

  /* ── FILE UPLOADER ── */
  [data-testid="stFileUploader"] {
    background: var(--surface) !important; border: 1.5px dashed var(--border-md) !important;
    border-radius: var(--r-lg) !important; box-shadow: var(--sh-card) !important;
    transition: border-color 160ms var(--ease-out), box-shadow 160ms var(--ease-out) !important; }
  @media (hover: hover) and (pointer: fine) {
    [data-testid="stFileUploader"]:hover {
      border-color: var(--accent) !important;
      box-shadow: 0 0 0 3px rgba(53,116,168,0.08), var(--sh-card) !important; } }
  [data-testid="stFileUploader"] label {
    color: #94A3B8 !important; font-family: var(--nx-mono) !important;
    font-size: 10px !important; letter-spacing: 1.5px !important; text-transform: uppercase !important; }
  [data-testid="stFileUploader"] span,
  [data-testid="stFileUploader"] p { color: var(--text-primary) !important; }
  [data-testid="stFileUploader"] section { background: var(--surface) !important; border: none !important; }
  [data-testid="stFileUploaderDropzone"] button {
    background: #F8FAFC !important; border: 1px solid var(--border) !important;
    border-radius: var(--r-sm) !important; color: #64748B !important; font-size: 12px !important; }
  [data-testid="stFileUploaderFile"] {
    background: #F8FAFC !important; border: 1px solid var(--border) !important;
    border-radius: var(--r-sm) !important; box-shadow: 0 1px 2px rgba(0,0,0,0.04) !important; }
  [data-testid="stFileUploaderFileName"] { color: var(--text-primary) !important; font-size: 12px !important; }
  [data-testid="stFileUploaderFileData"] {
    color: #94A3B8 !important; font-family: var(--nx-mono) !important; font-size: 10px !important; }

  /* ── TABS ── */
  .stTabs [data-baseweb="tab-list"] {
    background: transparent !important; border-radius: 0 !important;
    padding: 0 !important; border: none !important; box-shadow: none !important;
    border-bottom: 1px solid var(--border) !important; gap: 0 !important; }
  .stTabs [data-baseweb="tab"] {
    color: #94A3B8 !important; border-radius: 0 !important; font-weight: 400 !important;
    font-size: 13px !important; padding: 10px 18px !important;
    font-family: var(--nx-sans) !important; border-bottom: 2px solid transparent !important;
    background: transparent !important;
    transition: color 120ms var(--ease-out), border-color 120ms var(--ease-out) !important; }
  @media (hover: hover) and (pointer: fine) {
    .stTabs [data-baseweb="tab"]:hover { color: var(--text-primary) !important; } }
  .stTabs [aria-selected="true"] {
    color: var(--text-primary) !important; font-weight: 500 !important;
    border-bottom-color: var(--accent) !important;
    background: transparent !important; box-shadow: none !important; }

  .stProgress>div>div { background: var(--accent) !important; border-radius: 4px !important; }
  .stProgress { background: var(--border) !important; border-radius: 4px !important; }

  /* ── SLIDERS — CSS + attr selector para overridear inline styles de Streamlit ── */
  [data-testid="stSlider"] [role="slider"],
  [data-testid="stSlider"] [style*="rgb(255, 75, 75)"],
  [data-testid="stSlider"] [style*="rgb(255,75,75)"],
  [data-testid="stSlider"] [style*="rgb(255, 43, 43)"] {
    background: #3BADE8 !important;
    background-color: #3BADE8 !important;
    border-color: #3BADE8 !important;
    box-shadow: 0 0 0 4px rgba(59,173,232,0.2) !important; }
  [data-testid="stSlider"] [data-baseweb="slider"] > div > div:nth-child(2),
  [data-testid="stSlider"] [data-baseweb="slider"] > div > div:nth-child(3),
  [data-testid="stSlider"] [data-baseweb="slider"] > div > div:nth-child(4) > div {
    background: #3BADE8 !important; border-color: #3BADE8 !important; }
  [data-testid="stSlider"] p { color: #3BADE8 !important;
    font-family: var(--nx-mono) !important; font-size: 11px !important; }
  [data-testid="stSlider"] > label { color: rgba(240,246,252,0.88) !important;
    font-size: 12px !important; font-family: var(--nx-sans) !important; }

  /* ── METRIC CARDS — animate-skill: -4px lift, stronger hover shadow ── */
  .metric-card {
    background: var(--surface); border: 1px solid var(--border);
    border-top: 2px solid var(--accent); border-radius: var(--r-lg);
    padding: 16px 18px; text-align: left; box-shadow: var(--sh-card);
    transition: transform 200ms var(--ease-out), box-shadow 200ms var(--ease-out); }
  @media (hover: hover) and (pointer: fine) {
    .metric-card:hover {
      transform: translateY(-4px);
      box-shadow: 0 0 0 1px rgba(13,27,42,0.06), 0 12px 24px rgba(13,27,42,0.1), 0 24px 48px rgba(13,27,42,0.08); } }
  .metric-card h2 { font-family: var(--nx-sans); font-size: 1.9rem; margin: 0;
    font-weight: 700; color: var(--text-primary) !important;
    letter-spacing: -0.5px; font-variant-numeric: tabular-nums; }
  .metric-card p { margin: 0 0 4px; font-size: 10px; letter-spacing: 1.5px;
    text-transform: uppercase; font-family: var(--nx-mono); color: #94A3B8 !important; }

  /* ── ROW ITEMS — animate-skill: translateY(-4px) on cards ── */
  .nx-row { display: flex; align-items: flex-start; gap: 14px; padding: 14px 16px;
    border-bottom: 1px solid #EEF2F7; border-radius: var(--r-md);
    transition: background 140ms var(--ease-out),
                box-shadow 200ms var(--ease-out),
                transform 200ms var(--ease-out); }
  .nx-row:last-child { border-bottom: none; }
  @media (hover: hover) and (pointer: fine) {
    .nx-row:hover {
      background: var(--surface);
      box-shadow: 0 0 0 1px rgba(13,27,42,0.05), 0 4px 12px rgba(13,27,42,0.07), 0 12px 32px rgba(13,27,42,0.06);
      transform: translateY(-3px); }
    .nx-row:hover .nx-chev { color: var(--accent); } }
  /* frontend-design: min-width:0 on flex children with truncation */
  .nx-rank { font-family: var(--nx-mono); font-size: 11px; color: #CBD5E1;
    padding-top: 7px; min-width: 18px; flex-shrink: 0;
    font-variant-numeric: tabular-nums; }
  .nx-body { flex: 1; min-width: 0; }
  .nx-name { font-family: var(--nx-sans); font-size: 15px; font-weight: 600;
    color: var(--text-primary); letter-spacing: -0.2px; text-wrap: balance; }
  .nx-meta { font-family: var(--nx-mono); font-size: 11px; color: #94A3B8;
    letter-spacing: 0.2px; margin-bottom: 6px; font-variant-numeric: tabular-nums;
    overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
  .nx-desc { font-size: 13px; color: #475569; line-height: 1.55; text-wrap: pretty; }
  .nx-chev { padding-top: 9px; color: #CBD5E1;
    transition: color 140ms var(--ease-out); flex-shrink: 0; }
  .nx-section-label { display: flex; align-items: center; gap: 12px; margin: 10px 0 6px; }
  .nx-section-label span { font-family: var(--nx-mono); font-size: 10.5px; letter-spacing: 2px;
    text-transform: uppercase; color: #94A3B8; white-space: nowrap; }
  .nx-section-label .nx-rule { flex: 1; height: 1px; background: var(--border); }

  /* ── EMPTY STATE — frontend-design: invitation to act, not apology ── */
  .nx-empty {
    display: flex; flex-direction: column; align-items: center;
    justify-content: center; padding: 48px 24px; text-align: center;
    border-radius: var(--r-lg); background: var(--surface);
    border: 1.5px dashed var(--border-md); box-shadow: var(--sh-card); }
  .nx-empty-icon { font-size: 32px; margin-bottom: 16px;
    color: var(--border-md); line-height: 1; }
  .nx-empty-title { font-size: 15px; font-weight: 600; color: var(--text-primary);
    margin: 0 0 6px; letter-spacing: -0.2px; }
  .nx-empty-msg { font-size: 13px; color: var(--text-secondary);
    margin: 0 0 20px; line-height: 1.55; max-width: 300px; }

  /* ── BADGES ── */
  .badge-alto,.badge-prov-a   { background: #DCFCE7; color: #15803D; }
  .badge-medio,.badge-prov-c  { background: #FEF9C3; color: #854D0E; }
  .badge-bajo                 { background: #FEE2E2; color: #991B1B; }
  .badge-prov-b               { background: #D1FAE5; color: #065F46; }
  .badge-alto,.badge-medio,.badge-bajo,.badge-prov-a,.badge-prov-b,.badge-prov-c {
    font-family: var(--nx-mono); font-size: 10px; letter-spacing: 0.8px;
    text-transform: uppercase; padding: 3px 10px; border-radius: 20px; font-weight: 600; }

  /* ── ALERTS — semantic left-border ── */
  [data-testid="stAlertContainer"] {
    background: var(--surface) !important; border: 1px solid var(--border) !important;
    border-radius: var(--r-md) !important; padding: 12px 16px !important;
    box-shadow: var(--sh-card) !important; }
  [data-testid="stAlertContainer"] p { color: #475569 !important; font-size: 13px !important; }
  div[data-testid="stAlertContainer"]:has([data-testid="stNotificationContentInfo"]) {
    border-left: 3px solid #3574A8 !important; }
  div[data-testid="stAlertContainer"]:has([data-testid="stNotificationContentSuccess"]) {
    border-left: 3px solid #22C55E !important; }
  div[data-testid="stAlertContainer"]:has([data-testid="stNotificationContentWarning"]) {
    border-left: 3px solid #F59E0B !important; }
  div[data-testid="stAlertContainer"]:has([data-testid="stNotificationContentError"]) {
    border-left: 3px solid #EF4444 !important; }

  /* ── TYPOGRAPHY — frontend-design + web-design-guidelines ── */
  hr { border-color: var(--border) !important; }
  h1 { color: var(--text-primary) !important; font-family: var(--nx-sans) !important;
    letter-spacing: -0.5px; font-size: 1.6rem !important; font-weight: 700 !important;
    text-wrap: balance; }
  h2 { color: var(--text-primary) !important; font-family: var(--nx-sans) !important;
    letter-spacing: -0.3px; font-size: 1.25rem !important; font-weight: 600 !important;
    text-wrap: balance; }
  h3,h4 { color: var(--text-primary) !important; font-family: var(--nx-sans) !important;
    letter-spacing: -0.2px; text-wrap: balance; }
  p { text-wrap: pretty; }

  /* ── DATA TABLES — Linear tabular-nums ── */
  [data-testid="stDataFrame"] {
    border: 1px solid var(--border) !important; border-radius: var(--r-lg) !important;
    box-shadow: var(--sh-card) !important; }
  [data-testid="stDataFrame"] td,[data-testid="stDataFrame"] th {
    font-variant-numeric: tabular-nums !important; }

  /* ── FORM ── */
  .stForm {
    background: var(--surface) !important; border: 1px solid var(--border) !important;
    border-radius: var(--r-lg) !important; padding: 2rem !important;
    box-shadow: var(--sh-card) !important; }

  /* ── LOGO — glow ring ── */
  .logo-icon {
    width: 48px; height: 48px; border-radius: var(--r-lg); background: #1A2736;
    display: flex; align-items: center; justify-content: center;
    font-size: 18px; font-weight: 700; color: #F0F6FC;
    box-shadow: 0 0 0 3px rgba(53,116,168,0.28), 0 4px 14px rgba(13,27,42,0.3); }

  /* ── SKELETON — Stripe/Linear structural skeletons ── */
  @keyframes nx-shimmer {
    0%   { background-position: -400px 0; }
    100% { background-position:  400px 0; }
  }
  .nx-skeleton {
    border-radius: var(--r-sm);
    background: linear-gradient(90deg, #E2E8F0 25%, #F1F5F9 50%, #E2E8F0 75%);
    background-size: 800px 100%;
    animation: nx-shimmer 1.4s ease-in-out infinite; }

  /* ── STAGGER — animate-skill: enter 200ms, exit 150ms (~75%) ── */
  /* motion-principles: SaaS dashboard = Emil primary → fast, purposeful */
  .nx-row:nth-child(1) { animation: nx-fadein 200ms var(--ease-out) both; }
  .nx-row:nth-child(2) { animation: nx-fadein 200ms 40ms  var(--ease-out) both; }
  .nx-row:nth-child(3) { animation: nx-fadein 200ms 80ms  var(--ease-out) both; }
  .nx-row:nth-child(4) { animation: nx-fadein 200ms 120ms var(--ease-out) both; }
  .nx-row:nth-child(5) { animation: nx-fadein 200ms 160ms var(--ease-out) both; }
  @keyframes nx-fadein {
    from { opacity: 0; transform: translateY(6px); }
    to   { opacity: 1; transform: translateY(0); }
  }
  /* animate-skill: GPU-only (transform + opacity). NEVER layout props */
  /* web-design-guidelines + Linear + Emil: prefers-reduced-motion mandatory */
  @media (prefers-reduced-motion: reduce) {
    .nx-row { animation: none !important; }
    .nx-skeleton { animation: none !important; background: #E2E8F0; }
    * { transition-duration: 0.01ms !important; } }
</style>
<script>
(function() {
  var C = "#3BADE8";
  function fix() {
    document.querySelectorAll('[data-testid="stSlider"]').forEach(function(s) {
      s.querySelectorAll("div,span").forEach(function(el) {
        var bg = (el.style.backgroundColor||"")+(el.style.background||"");
        if(bg.indexOf("255, 75")>-1||bg.indexOf("255,75")>-1||
           bg.indexOf("255, 43")>-1||bg.indexOf("ff4b4b")>-1){
          el.style.setProperty("background-color",C,"important");
          el.style.setProperty("background",C,"important");
          el.style.setProperty("border-color",C,"important");
        }
      });
    });
  }
  var n=0,iv=setInterval(function(){fix();if(++n>30)clearInterval(iv);},300);
  window.addEventListener("load",fix);
})();
</script>
""", unsafe_allow_html=True)

# â”€â”€ LOGIN â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def mostrar_login():
    col1, col2, col3 = st.columns([1, 1.2, 1])
    with col2:
        st.markdown("""
        <div style="text-align:center;padding:2rem 0 1.5rem;">
          <div style="width:64px;height:64px;border-radius:16px;
               background:linear-gradient(135deg,#4A90B8,#2E3A45);
               display:inline-flex;align-items:center;justify-content:center;
               font-size:24px;font-weight:800;color:#FFFFFF;margin-bottom:16px;
               box-shadow:0 6px 20px rgba(46,58,69,0.25);">NX</div>
          <h2 style="color:#2E3A45;margin:0;font-size:1.8rem;font-weight:600;font-family:'Source Serif 4',serif;">Nexora</h2>
          <p style="color:#3A7CA5;margin:4px 0 0;font-size:12px;letter-spacing:3px;font-family:'IBM Plex Mono',monospace;text-transform:uppercase;">Analiza &middot; Decide &middot; Avanza</p>
          <p style="color:#9CA8B0;margin:4px 0 0;font-size:11px;letter-spacing:1px;">POWERED BY CLAUDE AI</p>
        </div>""", unsafe_allow_html=True)

name, authentication_status, username = authenticator.login(
    fields={"Form name": "Iniciar sesi\u00f3n", "Username": "Usuario", "Password": "Contrase\u00f1a", "Login": "Entrar"}
)
if authentication_status == False:
    mostrar_login(); st.error("Usuario o contrase\u00f1a incorrectos"); st.stop()
if authentication_status is None:
    mostrar_login(); st.info("Ingresa tus credenciales para acceder"); st.stop()

# â”€â”€ API â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
try:
    api_key_segura = st.secrets["CLAUDE_API_KEY"]
except Exception:
    st.error("No se encontr\u00f3 la API Key en los Secrets de Streamlit Cloud."); st.stop()
client = anthropic.Anthropic(api_key=api_key_segura)

# â”€â”€ SUPABASE (HISTORIAL) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
try:
    supabase = create_client(st.secrets["SUPABASE_URL"], st.secrets["SUPABASE_KEY"])
    SUPABASE_OK = True
except Exception:
    supabase = None
    SUPABASE_OK = False

TABLAS_HISTORIAL = {
    "cvs": "historial_cvs",
    "proveedores": "historial_proveedores",
    "propuestas": "historial_propuestas",
}

def guardar_historial(modulo, titulo, datos_json):
    """Guarda un registro de historial para el usuario actual.
    modulo: 'cvs' | 'proveedores' | 'propuestas'
    titulo: texto corto descriptivo (ej. '24 CVs - Analista de Datos')
    datos_json: dict serializable a JSON con todo lo necesario para restaurar la vista."""
    if not SUPABASE_OK:
        return
    tabla = TABLAS_HISTORIAL.get(modulo)
    if not tabla:
        return
    try:
        supabase.table(tabla).insert({
            "usuario": username,
            "titulo": titulo,
            "datos_json": datos_json,
        }).execute()
    except Exception as e:
        st.toast(f"No se pudo guardar en el historial: {e}", icon="\u26a0\ufe0f")

def listar_historial(modulo):
    """Devuelve la lista de registros de historial del usuario actual para un modulo,
    ordenados del mas reciente al mas antiguo. Cada item es un dict con id, fecha, titulo."""
    if not SUPABASE_OK:
        return []
    tabla = TABLAS_HISTORIAL.get(modulo)
    if not tabla:
        return []
    try:
        res = (supabase.table(tabla)
               .select("id, fecha, titulo")
               .eq("usuario", username)
               .order("fecha", desc=True)
               .execute())
        return res.data or []
    except Exception:
        return []

def cargar_historial_item(modulo, item_id):
    """Recupera el datos_json completo de un item especifico del historial."""
    if not SUPABASE_OK:
        return None
    tabla = TABLAS_HISTORIAL.get(modulo)
    if not tabla:
        return None
    try:
        res = (supabase.table(tabla)
               .select("datos_json")
               .eq("id", item_id)
               .eq("usuario", username)
               .single()
               .execute())
        return res.data.get("datos_json") if res.data else None
    except Exception:
        return None

def borrar_historial_item(modulo, item_id):
    """Elimina un item del historial del usuario actual."""
    if not SUPABASE_OK:
        return
    tabla = TABLAS_HISTORIAL.get(modulo)
    if not tabla:
        return
    try:
        supabase.table(tabla).delete().eq("id", item_id).eq("usuario", username).execute()
    except Exception as e:
        st.toast(f"No se pudo eliminar: {e}", icon="\u26a0\ufe0f")

def df_a_json_seguro(df):
    """Convierte un DataFrame a lista de dicts apta para jsonb de Supabase.
    Usa to_json/from_json para que NaN -> null y tipos numpy -> tipos nativos."""
    import json as _json
    return _json.loads(df.to_json(orient="records"))

def fecha_relativa(fecha_iso):
    """Convierte un timestamp ISO de Supabase a texto relativo corto: 'hoy', 'ayer', '12 jun'."""
    try:
        fecha = datetime.fromisoformat(fecha_iso.replace("Z", "+00:00"))
        ahora = datetime.now(timezone.utc)
        dias = (ahora.date() - fecha.date()).days
        if dias == 0: return "hoy"
        if dias == 1: return "ayer"
        if dias < 7:  return f"hace {dias}d"
        meses = ["ene","feb","mar","abr","may","jun","jul","ago","sep","oct","nov","dic"]
        return f"{fecha.day} {meses[fecha.month-1]}"
    except Exception:
        return ""

# â”€â”€ SESSION STATE â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
if "df_candidatos"   not in st.session_state: st.session_state.df_candidatos   = None
if "df_proveedores"  not in st.session_state: st.session_state.df_proveedores  = None
if "df_propuestas"   not in st.session_state: st.session_state.df_propuestas   = None
if "proveedores_web" not in st.session_state: st.session_state.proveedores_web = []
if "modulo_activo"   not in st.session_state: st.session_state.modulo_activo   = "cvs"
if "historial_item_activo" not in st.session_state: st.session_state.historial_item_activo = None

# â”€â”€ SIDEBAR â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
with st.sidebar:
    st.markdown(f"""
    <div style="padding:16px 0 10px;">
      <div style="display:flex;align-items:center;gap:12px;margin-bottom:6px;">
        <div style="width:44px;height:44px;border-radius:11px;flex-shrink:0;
             background:linear-gradient(135deg,#5B9EC9 0%,#2E6B9E 100%);
             display:flex;align-items:center;justify-content:center;
             font-size:15px;font-weight:800;color:#FFFFFF;
             box-shadow:0 0 0 2px rgba(91,158,201,0.35),0 4px 14px rgba(0,0,0,0.40);">NX</div>
        <div>
          <div class="nx-sb-primary" style="font-size:16px;line-height:1.2;letter-spacing:-0.2px;">Nexora</div>
          <div class="nx-sb-sub" style="font-size:9px;letter-spacing:2.5px;font-family:monospace;text-transform:uppercase;margin-top:3px;">Analiza &middot; Decide &middot; Avanza</div>
        </div>
      </div>
      <div class="nx-sb-muted" style="font-size:11.5px;margin-top:6px;padding-left:2px;">
        Bienvenido, <strong class="nx-sb-name">{name}</strong>
      </div>
    </div>""", unsafe_allow_html=True)
    st.divider()

    SIDEBAR_LABEL = "font-family:var(--nx-mono);font-size:10px;letter-spacing:1.8px;text-transform:uppercase;color:rgba(240,246,252,0.55);margin-bottom:10px;"

    _opts_mod = ["An\u00e1lisis de CVs", "An\u00e1lisis de Proveedores",
                 "An\u00e1lisis de Contratos", "Generador de RFQ", "An\u00e1lisis de Facturas"]
    _map_mod  = {"An\u00e1lisis de CVs": "cvs", "An\u00e1lisis de Proveedores": "proveedores",
                 "An\u00e1lisis de Contratos": "contratos", "Generador de RFQ": "rfq",
                 "An\u00e1lisis de Facturas": "facturas"}
    _idx_mod  = list(_map_mod.values()).index(st.session_state.modulo_activo) \
                if st.session_state.modulo_activo in _map_mod.values() else 0
    modulo = st.radio("M\u00f3dulo activo", _opts_mod, index=_idx_mod)
    st.session_state.modulo_activo = _map_mod[modulo]
    st.divider()

    # â”€â”€ HISTORIAL â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    if st.session_state.modulo_activo == "cvs":
        modulos_historial = [("cvs", "CVs")]
    elif st.session_state.modulo_activo == "proveedores":
        modulos_historial = [("proveedores", "B\u00fasquedas"), ("propuestas", "Propuestas")]
    else:
        modulos_historial = []

    st.markdown(f"<p style='{SIDEBAR_LABEL}'>Historial</p>", unsafe_allow_html=True)

    if not modulos_historial:
        st.markdown('<div class="nx-hist-empty">Sin historial para este módulo</div>', unsafe_allow_html=True)
    elif not SUPABASE_OK:
        st.markdown('<div class="nx-hist-empty">Historial no disponible</div>', unsafe_allow_html=True)
    else:
        hist_total = 0
        for mod_key, mod_label in modulos_historial:
            items = listar_historial(mod_key)
            hist_total += len(items)
            if not items:
                continue
            if len(modulos_historial) > 1:
                st.markdown(f'<div style="font-family:var(--nx-mono);font-size:9px;letter-spacing:1px;color:#B4B2A9;margin:6px 0 2px;text-transform:uppercase;">{mod_label}</div>', unsafe_allow_html=True)
            for item in items:
                item_id = item["id"]
                titulo  = item.get("titulo", "Sin t\u00edtulo")
                fecha_r = fecha_relativa(item.get("fecha", ""))
                es_activo = (st.session_state.historial_item_activo == (mod_key, item_id))
                clase = "nx-hist-item activo" if es_activo else "nx-hist-item"

                c1, c2 = st.columns([10, 1], gap="small")
                with c1:
                    st.markdown(f'<div class="{clase}"><span class="nx-hist-titulo">{esc(titulo)}</span><span class="nx-hist-fecha">{esc(fecha_r)}</span></div>', unsafe_allow_html=True)
                    if st.button("abrir", key=f"hist_open_{mod_key}_{item_id}", help=titulo):
                        datos = cargar_historial_item(mod_key, item_id)
                        if datos is not None:
                            if mod_key == "cvs":
                                st.session_state.df_candidatos = pd.DataFrame(datos)
                                st.session_state.modulo_activo = "cvs"
                            elif mod_key == "proveedores":
                                st.session_state.proveedores_web = datos.get("proveedores_web", [])
                                if datos.get("df_proveedores") is not None:
                                    st.session_state.df_proveedores = pd.DataFrame(datos["df_proveedores"])
                                st.session_state.modulo_activo = "proveedores"
                            elif mod_key == "propuestas":
                                st.session_state.df_propuestas = pd.DataFrame(datos)
                                if st.session_state.df_proveedores is None:
                                    st.session_state.df_proveedores = pd.DataFrame(datos)
                                else:
                                    st.session_state.df_proveedores = pd.concat(
                                        [st.session_state.df_proveedores, pd.DataFrame(datos)], ignore_index=True)
                                st.session_state.modulo_activo = "proveedores"
                            st.session_state.historial_item_activo = (mod_key, item_id)
                            st.rerun()
                with c2:
                    st.markdown('<div class="nx-hist-btn" style="margin-top:2px;">', unsafe_allow_html=True)
                    if st.button("\u00d7", key=f"hist_del_{mod_key}_{item_id}", help="Eliminar del historial"):
                        borrar_historial_item(mod_key, item_id)
                        if st.session_state.historial_item_activo == (mod_key, item_id):
                            st.session_state.historial_item_activo = None
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)
        if hist_total == 0:
            st.markdown('<div class="nx-hist-empty">A\u00fan no hay b\u00fasquedas guardadas</div>', unsafe_allow_html=True)

    st.divider()

    if st.session_state.modulo_activo == "cvs":
        st.markdown(f"<p style='{SIDEBAR_LABEL}'>Configurar puesto</p>", unsafe_allow_html=True)
        puesto          = st.text_input("Nombre del puesto", placeholder="Ej: Analista de Datos")
        experiencia_min = st.slider("\U0001f4c5 A\u00f1os m\u00ednimos de experiencia", 0, 20, 2)
        educacion_req   = st.selectbox("Educaci\u00f3n m\u00ednima", ["Cualquiera","T\u00e9cnico","Bachiller","Licenciatura","Maestr\u00eda","Doctorado"])
        habilidades_req = st.text_area("Habilidades requeridas (una por l\u00ednea)", placeholder="Python\nExcel\nSQL")
        idioma_req      = st.selectbox("Idioma requerido", ["No requerido","Ingl\u00e9s","Ingl\u00e9s avanzado","Portugu\u00e9s","Franc\u00e9s"])
        st.divider()
        st.markdown(f"<p style='{SIDEBAR_LABEL}'>Pesos de puntuaci\u00f3n</p>", unsafe_allow_html=True)
        peso_exp = st.slider("Experiencia", 0, 100, 35)
        peso_edu = st.slider("Educaci\u00f3n",   0, 100, 25)
        peso_hab = st.slider("Habilidades", 0, 100, 30)
        peso_idi = st.slider("Idiomas",     0, 100, 10)
    elif st.session_state.modulo_activo == "proveedores":
        st.markdown(f"<p style='{SIDEBAR_LABEL}'>Configurar b\u00fasqueda</p>", unsafe_allow_html=True)
        pais_busqueda   = st.text_input("Pa\u00eds o regi\u00f3n", placeholder="Ej: Per\u00fa, LATAM, Espa\u00f1a")
        rubro_busqueda  = st.text_input("Rubro o industria", placeholder="Ej: Software, Log\u00edstica")
        presupuesto_ref = st.selectbox("Presupuesto referencial", ["No especificado","< $10,000","$10,000 - $50,000","$50,000 - $200,000","> $200,000"])
        cert_requeridas = st.text_area("Certificaciones requeridas (una por l\u00ednea)", placeholder="ISO 9001\nAWS Certified")
        cobertura_req   = st.selectbox("Cobertura geogr\u00e1fica m\u00ednima", ["Local","Nacional","Regional LATAM","Internacional"])
        st.divider()
        st.markdown(f"<p style='{SIDEBAR_LABEL}'>Pesos de evaluaci\u00f3n</p>", unsafe_allow_html=True)
        ppeso_precio = st.slider("Precio/Condiciones", 0, 100, 30)
        ppeso_cert   = st.slider("Certificaciones",    0, 100, 25)
        ppeso_rep    = st.slider("Reputaci\u00f3n",         0, 100, 25)
        ppeso_cob    = st.slider("Cobertura",          0, 100, 20)
        st.divider()
        st.markdown(f"<p style='{SIDEBAR_LABEL}'>Matriz econ\u00f3mica</p>", unsafe_allow_html=True)
        peso_matriz_costo = st.slider("Costos de la propuesta", 0, 100, 70)
        peso_matriz_pago  = st.slider("Condici\u00f3n de pago",   0, 100, 30)

    st.divider()
    authenticator.logout("Cerrar sesi\u00f3n", "sidebar")

# â”€â”€ HEADER â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
st.markdown("""
<div style="display:flex;align-items:baseline;justify-content:space-between;
     padding-bottom:14px;border-bottom:1px solid #E5DFD3;margin-bottom:8px;">
  <div style="display:flex;align-items:center;gap:12px;">
    <div class="logo-icon" style="width:40px;height:40px;font-size:15px;">NX</div>
    <span style="font-family:var(--nx-serif);font-size:1.6rem;font-weight:600;color:#2E3A45;">Nexora</span>
  </div>
  <span style="font-family:var(--nx-mono);font-size:10px;letter-spacing:2.5px;color:#9CA8B0;text-transform:uppercase;">Analiza &middot; Decide &middot; Avanza</span>
</div>
""", unsafe_allow_html=True)

# â”€â”€ HELPERS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def limpiar_json(texto):
    texto = texto.strip()
    if "```" in texto:
        partes = texto.split("```")
        for p in partes:
            p = p.strip()
            if p.startswith("json"): p = p[4:].strip()
            if p.startswith("{") or p.startswith("["): texto = p; break
    inicio = texto.find("{"); fin = texto.rfind("}") + 1
    if inicio == -1 or fin == 0: raise ValueError("No se encontro JSON en la respuesta")
    return json.loads(texto[inicio:fin])

def safe_float(val):
    try: return float(str(val).replace(",", "."))
    except: return 0.0

def render_dial(puntaje, track_color, arc_color, size=56):
    """Genera un dial circular SVG que representa un puntaje 0-10.
    track_color: color del fondo del anillo (claro)
    arc_color: color del arco que representa el puntaje (oscuro)"""
    try:
        p = max(0.0, min(10.0, float(puntaje)))
    except (TypeError, ValueError):
        p = 0.0
    r = 22
    circunferencia = 2 * 3.14159265 * r
    relleno = round(circunferencia * (p / 10), 1)
    label = str(int(p)) if p == int(p) else f"{round(p,1):.1f}"
    return (
        f'<svg width="{size}" height="{size}" viewBox="0 0 56 56" style="flex-shrink:0;">'
        f'<circle cx="28" cy="28" r="{r}" fill="none" stroke="{track_color}" stroke-width="4"/>'
        f'<circle cx="28" cy="28" r="{r}" fill="none" stroke="{arc_color}" stroke-width="4" '
        f'stroke-linecap="round" stroke-dasharray="{relleno} {circunferencia:.1f}" '
        f'transform="rotate(-90 28 28)"/>'
        f'<text x="28" y="33" text-anchor="middle" font-family="\'Source Serif 4\',serif" '
        f'font-size="18" font-weight="600" fill="#2E3A45">{label}</text>'
        f'</svg>'
    )

def dial_colores_cv(nivel):
    """Devuelve (track_color, arc_color) segun el nivel de potencial de un CV."""
    if nivel == "Alto":  return "#E6F4E9", "#3D7A4D"
    if nivel == "Medio": return "#FBF1DD", "#A8762E"
    return "#FAE6E6", "#B3504F"

def dial_colores_prov(puntaje):
    """Devuelve (track_color, arc_color) segun el puntaje de recomendacion de un proveedor."""
    p = safe_float(puntaje)
    if p >= 8: return "#E1F5EE", "#0F6E56"
    if p >= 6: return "#EAF3DE", "#5E7D3C"
    return "#FBF1DD", "#A8762E"

def banner_archivos(cantidad, texto):
    """Banner de archivos cargados con numero en serif a la derecha."""
    n = f"{cantidad:02d}"
    st.markdown(f"""<div style="display:flex;align-items:center;gap:10px;background:#FFFFFF;
      border:1px solid #E5DFD3;border-left:2px solid #4A90B8;border-radius:8px;
      padding:10px 14px;margin-bottom:14px;">
      <i class="ti ti-info-circle" style="font-size:15px;color:#4A90B8;"></i>
      <span style="font-size:12px;color:#5F5E5A;">{texto}</span>
      <span style="margin-left:auto;font-family:var(--nx-serif);font-size:14px;font-weight:600;color:#2E3A45;">{n}</span>
    </div>""", unsafe_allow_html=True)

def progreso_label(label, actual, total):
    """Render 'label ---- actual/total' al estilo nx-section-label, con contador en serif."""
    st.markdown(f"""<div class="nx-section-label" style="margin-top:14px;">
      <span>{label}</span><div class="nx-rule"></div>
      <span style="font-family:var(--nx-serif);font-size:13px;font-weight:600;color:#4A90B8;white-space:nowrap;">{actual:02d} / {total:02d}</span>
    </div>""", unsafe_allow_html=True)

def titulo_modulo(titulo, subtitulo):
    """Titulo de modulo: serif grande + label mono debajo, estilo masthead."""
    st.markdown(f"""<div style="margin-bottom:18px;">
      <div style="font-family:var(--nx-serif);font-size:1.6rem;font-weight:600;color:#2E3A45;">{titulo}</div>
      <div style="font-family:var(--nx-mono);font-size:10px;letter-spacing:2px;text-transform:uppercase;color:#9CA8B0;margin-top:2px;">{subtitulo}</div>
    </div>""", unsafe_allow_html=True)

def preparar_texto_documento(texto, limite=18000):
    """Prepara el texto del documento para enviarlo a Claude.
    Si el documento es largo, antepone las secciones que contienen
    palabras clave criticas (garantia, tiempo de entrega, IGV, condiciones
    de pago, totales) para garantizar que no se pierdan por el limite
    de caracteres, sin sobrepasar nunca el limite total."""
    if len(texto) <= limite:
        return texto

    palabras_clave = [
        "GARANT", "TIEMPO DE ENTREGA", "PLAZO DE ENTREGA", "ENTREGA:",
        " IGV", "I.G.V", "IVA", "SUBTOTAL", "SUB TOTAL", " TOTAL",
        "CONDICION", "FORMA DE PAGO", "CREDITO", "CR\u00c9DITO",
        "VALIDEZ", "VIGENCIA", "PRECIO UNITARIO", "S/.", "$",
    ]

    lineas = texto.split("\n")
    indices_clave = set()
    for i, linea in enumerate(lineas):
        linea_upper = linea.upper()
        if any(pk in linea_upper for pk in palabras_clave):
            for j in range(max(0, i - 1), min(len(lineas), i + 3)):
                indices_clave.add(j)

    lineas_clave = [lineas[i] for i in sorted(indices_clave)]
    seccion_clave = "\n".join(lineas_clave)

    encabezado = (
        "\n\n--- INFORMACION ADICIONAL DETECTADA EN EL DOCUMENTO "
        "(condiciones comerciales, garantia, plazos, precios) ---\n"
    )

    # Reservar espacio fijo para la seccion clave (maximo 40% del limite)
    max_seccion = min(len(seccion_clave), int(limite * 0.4))
    seccion_clave = seccion_clave[:max_seccion]

    espacio_para_principal = limite - len(encabezado) - len(seccion_clave)
    espacio_para_principal = max(500, espacio_para_principal)

    texto_principal = texto[:espacio_para_principal]

    if seccion_clave:
        return texto_principal + encabezado + seccion_clave
    return texto[:limite]

def exportar_excel_cvs(df_exp):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_exp.to_excel(writer, index=False, sheet_name="Candidatos")
        ws = writer.sheets["Candidatos"]
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        header_fill = PatternFill(start_color="26215C", end_color="26215C", fill_type="solid")
        alto_fill   = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        medio_fill  = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
        bajo_fill   = PatternFill(start_color="FDECEA", end_color="FDECEA", fill_type="solid")
        par_fill    = PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid")
        blanco_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        borde = Border(
            left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin",  color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC")
        )
        anchos = {
            "Ranking":8,"Nombre":24,"Correo":30,"Telefono":16,
            "Educacion_Maxima":20,"Universidad":26,"Carrera":22,
            "Ultimo_Cargo":26,"Ultima_Empresa":24,"Experiencia_Anos":14,
            "Habilidades_Tecnicas":38,"Habilidades_Blandas":30,
            "Idiomas":16,"Certificaciones":26,"Puntaje":10,
            "Nivel_Potencial":16,"Justificacion":60,
            "Cumple_Requisitos":16,"Requisitos_Cumplidos":32,
            "Requisitos_Faltantes":32,"Archivo":30,
        }
        nombres_col = {
            "Ranking":"Ranking","Nombre":"Nombre","Correo":"Correo",
            "Telefono":"Tel\u00e9fono","Educacion_Maxima":"Educaci\u00f3n M\u00e1xima",
            "Universidad":"Universidad","Carrera":"Carrera",
            "Ultimo_Cargo":"\u00daltimo Cargo","Ultima_Empresa":"\u00daltima Empresa",
            "Experiencia_Anos":"A\u00f1os de Experiencia","Habilidades_Tecnicas":"Habilidades T\u00e9cnicas",
            "Habilidades_Blandas":"Habilidades Blandas","Idiomas":"Idiomas",
            "Certificaciones":"Certificaciones","Puntaje":"Puntaje",
            "Nivel_Potencial":"Nivel de Potencial","Justificacion":"Justificaci\u00f3n",
            "Cumple_Requisitos":"Cumple Requisitos","Requisitos_Cumplidos":"Requisitos Cumplidos",
            "Requisitos_Faltantes":"Requisitos Faltantes","Archivo":"Archivo",
        }
        for cn, col in enumerate(df_exp.columns, 1):
            c = ws.cell(row=1, column=cn)
            c.value     = nombres_col.get(col, col.replace("_"," "))
            c.font      = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            c.fill      = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = borde
            ws.column_dimensions[c.column_letter].width = anchos.get(col, 20)
        ws.row_dimensions[1].height = 36
        nivel_col_idx = next((ci for ci, col in enumerate(df_exp.columns, 1) if col == "Nivel_Potencial"), None)
        for rn, (_, row_data) in enumerate(df_exp.iterrows(), 2):
            nivel = str(row_data.get("Nivel_Potencial", ""))
            if nivel == "Alto":   fila_fill = alto_fill;  nivel_color = "1B5E20"
            elif nivel == "Medio": fila_fill = medio_fill; nivel_color = "E65100"
            elif nivel == "Bajo":  fila_fill = bajo_fill;  nivel_color = "B71C1C"
            else: fila_fill = par_fill if rn % 2 == 0 else blanco_fill; nivel_color = "000000"
            for cn in range(1, len(df_exp.columns) + 1):
                c = ws.cell(row=rn, column=cn)
                c.font      = Font(size=10, name="Calibri", color="1a1035")
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                c.border    = borde; c.fill = fila_fill
                if cn == nivel_col_idx:
                    c.font = Font(size=10, name="Calibri", bold=True, color=nivel_color)
                if df_exp.columns[cn-1] == "Puntaje":
                    c.font = Font(size=11, name="Calibri", bold=True, color="26215C")
                    c.alignment = Alignment(horizontal="center", vertical="center")
            max_lineas = 1
            for cn2 in range(1, len(df_exp.columns) + 1):
                valor = str(ws.cell(row=rn, column=cn2).value or "")
                col_name2 = df_exp.columns[cn2-1]
                ancho_col = anchos.get(col_name2, 20)
                chars_por_linea = max(int(ancho_col * 1.15), 12)
                lineas = max(1, -(-len(valor) // chars_por_linea))
                if lineas > max_lineas:
                    max_lineas = lineas
            ws.row_dimensions[rn].height = max(40, min(max_lineas * 16, 280))
        ws.freeze_panes = "A2"
    return output.getvalue()

def exportar_excel_proveedores(df_exp):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_exp.to_excel(writer, index=False, sheet_name="Proveedores")
        ws = writer.sheets["Proveedores"]
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        header_fill  = PatternFill(start_color="085041", end_color="085041", fill_type="solid")
        muy_rec_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        rec_fill     = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        viable_fill  = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
        norec_fill   = PatternFill(start_color="FDECEA", end_color="FDECEA", fill_type="solid")
        par_fill     = PatternFill(start_color="F0FFF4", end_color="F0FFF4", fill_type="solid")
        blanco_fill  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        borde = Border(
            left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin",  color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC")
        )
        anchos = {
            "nombre":26,"nombre_empresa":26,"descripcion":40,"sitio_web":30,
            "pais_sede":14,"cobertura":16,"anos_experiencia":14,"certificaciones":28,
            "productos_servicios":36,"rango_precio":20,"condiciones_comerciales":30,
            "tiempo_entrega":16,"clientes_referencia":30,"fortalezas":36,"debilidades":30,
            "puntaje_precio":12,"puntaje_certificaciones":16,"puntaje_reputacion":14,
            "puntaje_cobertura":14,"puntaje_recomendacion":16,"nivel_recomendacion":20,
            "justificacion":44,"razon_recomendacion":44,"cumple_certificaciones":18,
            "certificaciones_faltantes":30,"contacto":24,"Archivo":30,"Fuente":14,
        }
        nombres_col_prov = {
            "nombre":"Nombre","nombre_empresa":"Nombre Empresa",
            "descripcion":"Descripci\u00f3n","sitio_web":"Sitio Web",
            "pais_sede":"Pa\u00eds Sede","cobertura":"Cobertura",
            "anos_experiencia":"A\u00f1os de Experiencia","certificaciones":"Certificaciones",
            "productos_servicios":"Productos / Servicios","rango_precio":"Rango de Precio",
            "condiciones_comerciales":"Condiciones Comerciales","tiempo_entrega":"Tiempo de Entrega",
            "clientes_referencia":"Clientes de Referencia","fortalezas":"Fortalezas",
            "debilidades":"Debilidades","puntaje_precio":"Puntaje Precio",
            "puntaje_certificaciones":"Puntaje Certificaciones",
            "puntaje_reputacion":"Puntaje Reputaci\u00f3n","puntaje_cobertura":"Puntaje Cobertura",
            "puntaje_recomendacion":"Puntaje General","nivel_recomendacion":"Nivel de Recomendaci\u00f3n",
            "justificacion":"Justificaci\u00f3n","razon_recomendacion":"Raz\u00f3n de Recomendaci\u00f3n",
            "cumple_certificaciones":"Cumple Certificaciones",
            "certificaciones_faltantes":"Certificaciones Faltantes",
            "contacto":"Contacto","Archivo":"Archivo","Fuente":"Fuente",
        }
        for cn, col in enumerate(df_exp.columns, 1):
            c = ws.cell(row=1, column=cn)
            c.value     = nombres_col_prov.get(col, col.replace("_"," ").title())
            c.font      = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            c.fill      = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = borde
            ws.column_dimensions[c.column_letter].width = anchos.get(col, 20)
        ws.row_dimensions[1].height = 36
        nivel_col_idx = next((ci for ci, col in enumerate(df_exp.columns, 1) if col == "nivel_recomendacion"), None)
        for rn, (_, row_data) in enumerate(df_exp.iterrows(), 2):
            nivel = str(row_data.get("nivel_recomendacion", ""))
            if "Muy" in nivel:        fila_fill = muy_rec_fill; nivel_color = "1B5E20"
            elif nivel == "Recomendado": fila_fill = rec_fill;  nivel_color = "0D47A1"
            elif "viable" in nivel.lower(): fila_fill = viable_fill; nivel_color = "E65100"
            elif "No" in nivel:       fila_fill = norec_fill;  nivel_color = "B71C1C"
            else: fila_fill = par_fill if rn % 2 == 0 else blanco_fill; nivel_color = "000000"
            for cn in range(1, len(df_exp.columns) + 1):
                c = ws.cell(row=rn, column=cn)
                c.font      = Font(size=10, name="Calibri", color="1a2e1a")
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                c.border    = borde; c.fill = fila_fill
                if cn == nivel_col_idx:
                    c.font = Font(size=10, name="Calibri", bold=True, color=nivel_color)
                if "puntaje" in df_exp.columns[cn-1]:
                    c.font = Font(size=11, name="Calibri", bold=True, color="085041")
                    c.alignment = Alignment(horizontal="center", vertical="center")
            max_lineas = 1
            for cn2 in range(1, len(df_exp.columns) + 1):
                valor = str(ws.cell(row=rn, column=cn2).value or "")
                col_name2 = df_exp.columns[cn2-1]
                ancho_col = anchos.get(col_name2, 20)
                chars_por_linea = max(int(ancho_col * 1.15), 12)
                lineas = max(1, -(-len(valor) // chars_por_linea))
                if lineas > max_lineas:
                    max_lineas = lineas
            ws.row_dimensions[rn].height = max(40, min(max_lineas * 16, 280))
        ws.freeze_panes = "A2"
    return output.getvalue()

def exportar_excel_propuestas(df_exp):
    """Excel enfocado en condiciones comerciales de propuestas: garantia,
    tiempo de entrega, precios con/sin IGV y condiciones de pago."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_exp.to_excel(writer, index=False, sheet_name="Propuestas")
        ws = writer.sheets["Propuestas"]
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        header_fill = PatternFill(start_color="2E3A45", end_color="2E3A45", fill_type="solid")
        par_fill    = PatternFill(start_color="F0F6FB", end_color="F0F6FB", fill_type="solid")
        blanco_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        destaca_fill= PatternFill(start_color="EAF3E6", end_color="EAF3E6", fill_type="solid")
        borde = Border(
            left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin",  color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC")
        )

        anchos = {
            "nombre": 26, "Archivo": 30,
            "garantia": 32, "tiempo_entrega": 24,
            "precio_sin_igv": 18, "precio_con_igv": 18,
            "condiciones_pago": 36, "rango_precio": 20,
            "condiciones_comerciales": 30,
        }
        nombres_col = {
            "nombre": "Proveedor", "Archivo": "Archivo",
            "garantia": "Garant\u00eda",
            "tiempo_entrega": "Tiempo de Entrega",
            "precio_sin_igv": "Precio sin IGV",
            "precio_con_igv": "Precio con IGV",
            "condiciones_pago": "Condiciones de Pago",
            "rango_precio": "Rango de Precio",
            "condiciones_comerciales": "Condiciones Comerciales",
        }

        # Encabezados
        for cn, col in enumerate(df_exp.columns, 1):
            c = ws.cell(row=1, column=cn)
            c.value     = nombres_col.get(col, col.replace("_", " ").title())
            c.font      = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            c.fill      = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = borde
            ws.column_dimensions[c.column_letter].width = anchos.get(col, 22)
        ws.row_dimensions[1].height = 36

        # Columnas que se resaltan (las 4 solicitadas)
        cols_destacadas = {"garantia", "tiempo_entrega", "precio_sin_igv", "precio_con_igv", "condiciones_pago"}

        for rn, (_, row_data) in enumerate(df_exp.iterrows(), 2):
            fila_fill = par_fill if rn % 2 == 0 else blanco_fill
            for cn in range(1, len(df_exp.columns) + 1):
                col_name = df_exp.columns[cn-1]
                c = ws.cell(row=rn, column=cn)
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                c.border    = borde
                if col_name in cols_destacadas:
                    c.fill = destaca_fill
                    c.font = Font(size=10, name="Calibri", bold=True, color="2E5E3D")
                else:
                    c.fill = fila_fill
                    c.font = Font(size=10, name="Calibri", color="2E3A45")
                if col_name == "nombre":
                    c.font = Font(size=11, name="Calibri", bold=True, color="2E3A45")

            max_lineas = 1
            for cn2 in range(1, len(df_exp.columns) + 1):
                valor = str(ws.cell(row=rn, column=cn2).value or "")
                col_name2 = df_exp.columns[cn2-1]
                ancho_col = anchos.get(col_name2, 22)
                chars_por_linea = max(int(ancho_col * 1.15), 12)
                lineas = max(1, -(-len(valor) // chars_por_linea))
                if lineas > max_lineas:
                    max_lineas = lineas
            ws.row_dimensions[rn].height = max(40, min(max_lineas * 16, 200))

        ws.freeze_panes = "A2"
    return output.getvalue()



def exportar_excel_matriz_economica(df_propuestas, peso_costo, peso_pago, username,
                                      descripcion_compra="", comprador=""):
    """Genera un Excel estilo 'Matriz de Evaluacion Economica' (formato Yanbal),
    comparando hasta 5 proveedores lado a lado con formulas reales de Excel.

    df_propuestas: DataFrame con las propuestas analizadas (resultados_prov),
                    debe incluir columnas: nombre, item_descripcion, item_cantidad,
                    moneda, costo_unitario_num, costo_total_num, garantia,
                    tiempo_entrega, condiciones_pago, Archivo.
    peso_costo, peso_pago: enteros 0-100 (porcentajes) desde el sidebar.
    username: usuario logueado actual (campo "Usuario" del encabezado).
    descripcion_compra, comprador: textos opcionales para el encabezado.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.properties import PageSetupProperties

    NARANJA      = "E8703A"
    NARANJA_TXT  = "FFFFFF"
    GRIS_LABEL   = "2E3A45"
    AMARILLO     = "FFF2A6"
    VERDE_CLARO  = "D6EAD7"
    ROSADO       = "F4D6D6"
    AZUL_LINK    = "1F5C8B"
    BORDE_COLOR  = "D9D2C4"

    borde = Border(
        left=Side(style="thin", color=BORDE_COLOR), right=Side(style="thin", color=BORDE_COLOR),
        top=Side(style="thin", color=BORDE_COLOR), bottom=Side(style="thin", color=BORDE_COLOR),
    )
    fill_header  = PatternFill(start_color=NARANJA, end_color=NARANJA, fill_type="solid")
    fill_amarillo = PatternFill(start_color=AMARILLO, end_color=AMARILLO, fill_type="solid")
    fill_verde   = PatternFill(start_color=VERDE_CLARO, end_color=VERDE_CLARO, fill_type="solid")
    fill_rosado  = PatternFill(start_color=ROSADO, end_color=ROSADO, fill_type="solid")
    font_header  = Font(bold=True, color=NARANJA_TXT, size=11, name="Calibri")
    font_label   = Font(bold=True, color=GRIS_LABEL, size=10, name="Calibri")
    font_normal  = Font(size=10, name="Calibri", color=GRIS_LABEL)
    font_bold    = Font(bold=True, size=10, name="Calibri", color=GRIS_LABEL)

    df = df_propuestas.copy()
    # Maximo 5 proveedores (como en la plantilla Yanbal)
    df = df.head(5).reset_index(drop=True)
    n_prov = len(df)
    n_cols_total_max = n_prov  # columnas dinamicas: solo los proveedores reales

    col_nombre = "nombre" if "nombre" in df.columns else "Archivo"

    def gv(row, campo, default=""):
        val = row.get(campo, default)
        if val in [None, "", "No especifica"]:
            return default
        return val

    def gnum(row, campo, default=0.0):
        val = row.get(campo, default)
        try:
            f = float(val)
            if f != f:  # NaN check
                return default
            return f
        except (TypeError, ValueError):
            return default

    wb = Workbook()
    ws = wb.active
    ws.title = "Matriz Economica"

    # Ancho de columnas: A=labels de encabezado / Cant, B=Item, luego 3 columnas por proveedor
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    col_idx = 3  # columna C en adelante
    grupo_cols = []  # lista de (col_unitario, col_total, col_ofertas) por proveedor
    for i in range(n_cols_total_max):
        c_uni, c_tot, c_ofe = col_idx, col_idx + 1, col_idx + 2
        ws.column_dimensions[get_column_letter(c_uni)].width = 16
        ws.column_dimensions[get_column_letter(c_tot)].width = 16
        ws.column_dimensions[get_column_letter(c_ofe)].width = 14
        grupo_cols.append((c_uni, c_tot, c_ofe))
        col_idx += 3
    total_cols = col_idx - 1

    # â”€â”€ ENCABEZADO â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    fila = 1
    ws.cell(row=fila, column=1, value="MATRIZ EVALUACION ECONOMICA").font = Font(bold=True, size=13, color=GRIS_LABEL, name="Calibri")
    fila += 1
    fila += 1

    # Banda naranja "Proceso" -- merge sobre TODAS las columnas
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=total_cols)
    cell_proceso = ws.cell(row=fila, column=1, value="Proceso: A4 GESTION DE ADQUISICIONES")
    cell_proceso.fill = fill_header
    cell_proceso.font = font_header
    cell_proceso.alignment = Alignment(horizontal="center")
    for c in range(1, total_cols + 1):
        ws.cell(row=fila, column=c).fill = fill_header
    fila += 1
    fila += 1

    encabezado_items = [
        ("Descripcion de compra", descripcion_compra or "Analisis de propuestas de proveedores"),
        ("Usuario", username),
        ("Comprador", comprador or username),
        ("N\u00b0 PR", "-"),
        ("Fecha de solicitud de compra", datetime.now().strftime("%d/%m/%Y")),
        ("Fecha de entrega", "-"),
        ("Vigencia de la contrataci\u00f3n", "-"),
    ]
    for label, valor in encabezado_items:
        ws.cell(row=fila, column=1, value=label).font = font_label
        ws.cell(row=fila, column=2, value=str(valor)).font = font_normal
        fila += 1

    # "Proveedores invitados" lista hacia abajo en columna B
    ws.cell(row=fila, column=1, value="Proveedores invitados").font = font_label
    for i in range(n_prov):
        nombre_p = gv(df.iloc[i], col_nombre, f"Proveedor {i+1}")
        ws.cell(row=fila, column=2, value=str(nombre_p)).font = font_normal
        fila += 1
    fila += 1

    # â”€â”€ TABLA "PROVEEDORES" â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws.cell(row=fila, column=1, value="Proveedores")
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=fila, column=c)
        cell.fill = fill_header
        cell.font = font_header
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=total_cols)
    ws.cell(row=fila, column=1).alignment = Alignment(horizontal="center")
    fila += 1

    # Fila de nombres de proveedor (merge de 3 columnas cada uno)
    for i, (c_uni, c_tot, c_ofe) in enumerate(grupo_cols):
        nombre_p = gv(df.iloc[i], col_nombre, f"Proveedor {i+1}") if i < n_prov else "0"
        ws.merge_cells(start_row=fila, start_column=c_uni, end_row=fila, end_column=c_ofe)
        cell = ws.cell(row=fila, column=c_uni, value=str(nombre_p))
        cell.font = font_bold
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="FBE5D6", end_color="FBE5D6", fill_type="solid")
        for cc in range(c_uni, c_ofe + 1):
            ws.cell(row=fila, column=cc).border = borde
    fila += 1

    # Encabezados de columna: Cant | Item | Costo unitario | Costo total | Ofertas Adjuntas (x5)
    h_cant = ws.cell(row=fila, column=1, value="Cant")
    h_item = ws.cell(row=fila, column=2, value="Item")
    for cell in (h_cant, h_item):
        cell.font = font_bold
        cell.fill = PatternFill(start_color="FBE5D6", end_color="FBE5D6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde
    for (c_uni, c_tot, c_ofe) in grupo_cols:
        for c, txt in ((c_uni, "Costo unitario"), (c_tot, "Costo total"), (c_ofe, "Ofertas Adjuntas")):
            cell = ws.cell(row=fila, column=c, value=txt)
            cell.font = font_bold
            cell.fill = PatternFill(start_color="FBE5D6", end_color="FBE5D6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = borde
    fila += 1

    # Fila de Item (1 sola fila de producto, como en el ejemplo Yanbal)
    fila_item = fila
    item_desc = ""
    item_cant = 1
    for i in range(n_prov):
        d = gv(df.iloc[i], "item_descripcion", "")
        if d:
            item_desc = str(d)
            item_cant = int(gnum(df.iloc[i], "item_cantidad", 1)) or 1
            break
    if not item_desc:
        item_desc = "Producto/servicio cotizado"

    ws.cell(row=fila, column=1, value=item_cant).alignment = Alignment(horizontal="center")
    ws.cell(row=fila, column=1).font = font_normal
    ws.cell(row=fila, column=1).border = borde
    ws.cell(row=fila, column=2, value=item_desc).font = font_normal
    ws.cell(row=fila, column=2).border = borde
    ws.cell(row=fila, column=2).alignment = Alignment(wrap_text=True, vertical="center")

    costo_total_refs = []  # celdas de Costo total por proveedor, para la formula de TOTAL y de puntuacion
    for i, (c_uni, c_tot, c_ofe) in enumerate(grupo_cols):
        if i < n_prov:
            moneda = gv(df.iloc[i], "moneda", "S/.")
            simbolo = "$" if str(moneda).upper() == "USD" else "S/."
            costo_uni = gnum(df.iloc[i], "costo_unitario_num", 0.0)
            costo_tot = gnum(df.iloc[i], "costo_total_num", 0.0) or costo_uni

            cell_uni = ws.cell(row=fila, column=c_uni, value=costo_uni)
            cell_uni.number_format = f'"{simbolo}" #,##0.00'
            cell_uni.font = font_normal
            cell_uni.border = borde

            cell_tot = ws.cell(row=fila, column=c_tot)
            cell_tot.value = f"=+{get_column_letter(c_uni)}{fila}"
            cell_tot.number_format = f'"{simbolo}" #,##0.00'
            cell_tot.font = font_normal
            cell_tot.border = borde
            costo_total_refs.append((get_column_letter(c_tot), fila, costo_tot))

            archivo = gv(df.iloc[i], "Archivo", "")
            cell_ofe = ws.cell(row=fila, column=c_ofe, value=str(archivo))
            cell_ofe.font = Font(size=9, name="Calibri", color=AZUL_LINK)
            cell_ofe.border = borde
    fila += 1

    # 2 filas vacias adicionales (como en la plantilla, para items extra manuales)
    filas_vacias_inicio = fila
    for _ in range(2):
        ws.cell(row=fila, column=1).border = borde
        ws.cell(row=fila, column=2).border = borde
        for (c_uni, c_tot, c_ofe) in grupo_cols:
            for c in (c_uni, c_tot, c_ofe):
                cell = ws.cell(row=fila, column=c)
                cell.border = borde
                if c in [g[1] for g in grupo_cols]:
                    cell.number_format = '"S/." #,##0.00'
        fila += 1
    filas_vacias_fin = fila - 1

    # Fila TOTAL = SUMA(rango) por proveedor
    ws.cell(row=fila, column=2, value="TOTAL").font = font_bold
    ws.cell(row=fila, column=2).border = borde
    ws.cell(row=fila, column=1).border = borde

    colores_total = [fill_rosado, fill_amarillo, fill_verde, fill_amarillo, fill_verde]
    formula_total_refs = []
    for i, (c_uni, c_tot, c_ofe) in enumerate(grupo_cols):
        moneda = gv(df.iloc[i], "moneda", "S/.") if i < n_prov else "S/."
        simbolo = "$" if str(moneda).upper() == "USD" else "S/."
        letra_tot = get_column_letter(c_tot)
        formula = f"=SUM({letra_tot}{fila_item}:{letra_tot}{filas_vacias_fin})"
        cell = ws.cell(row=fila, column=c_tot, value=formula)
        cell.number_format = f'"{simbolo}" #,##0.00'
        cell.font = font_bold
        cell.fill = colores_total[i % len(colores_total)]
        cell.border = borde
        ws.cell(row=fila, column=c_uni).border = borde
        ws.cell(row=fila, column=c_ofe).border = borde
        formula_total_refs.append((letra_tot, fila))
    fila += 2

    # â”€â”€ TABLA "EVALUACION COMERCIAL" â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # Layout: columna B = Criterio, columna C = Peso, columnas D.. = Puntuacion por proveedor
    ancho_eval = 2 + n_cols_total_max  # B(criterio)+C(peso) + 1 col por proveedor
    ws.cell(row=fila, column=2, value="Evaluaci\u00f3n Comercial")
    ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=2 + ancho_eval - 1)
    cell = ws.cell(row=fila, column=2)
    cell.fill = fill_header
    cell.font = font_header
    cell.alignment = Alignment(horizontal="center")
    for c in range(2, 2 + ancho_eval):
        ws.cell(row=fila, column=c).fill = fill_header
    fila += 1

    h_crit = ws.cell(row=fila, column=2, value="Criterio de selecci\u00f3n")
    h_peso = ws.cell(row=fila, column=3, value="Peso")
    for cell in (h_crit, h_peso):
        cell.font = font_bold
        cell.fill = PatternFill(start_color="FBE5D6", end_color="FBE5D6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde
    for i in range(n_cols_total_max):
        col_punt = 4 + i
        nombre_p = gv(df.iloc[i], col_nombre, f"Proveedor {i+1}") if i < n_prov else "0"
        cell = ws.cell(row=fila, column=col_punt, value=str(nombre_p))
        cell.font = font_bold
        cell.fill = PatternFill(start_color="FBE5D6", end_color="FBE5D6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde
    fila += 1

    # Subfila "Puntuacion"
    for c in (2, 3):
        ws.cell(row=fila, column=c).border = borde
        ws.cell(row=fila, column=c).fill = PatternFill(start_color="FDF1E8", end_color="FDF1E8", fill_type="solid")
    for i in range(n_cols_total_max):
        col_punt = 4 + i
        cell = ws.cell(row=fila, column=col_punt, value="Puntuaci\u00f3n")
        cell.font = font_bold
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="FDF1E8", end_color="FDF1E8", fill_type="solid")
        cell.border = borde
    fila += 1

    # Fila "Costos de la propuesta" -- formula =mejor_costo/este_costo*5
    fila_costos = fila
    ws.cell(row=fila, column=2, value="Costos de la propuesta").font = font_normal
    ws.cell(row=fila, column=2).border = borde
    cell_peso_costo = ws.cell(row=fila, column=3, value=peso_costo / 100)
    cell_peso_costo.number_format = "0%"
    cell_peso_costo.font = font_normal
    cell_peso_costo.alignment = Alignment(horizontal="center")
    cell_peso_costo.border = borde

    # Determinar el mejor (menor) costo total entre los proveedores con costo > 0.
    # costo_total_refs trae el valor numerico real (para comparar); formula_total_refs
    # trae la celda (letra, fila_total) donde vive la formula SUMA correspondiente.
    costos_validos = [
        (formula_total_refs[i][0], formula_total_refs[i][1], costo_total_refs[i][2])
        for i in range(min(len(formula_total_refs), len(costo_total_refs)))
        if costo_total_refs[i][2] > 0
    ]
    if costos_validos:
        mejor_letra, mejor_fila, _ = min(costos_validos, key=lambda x: x[2])
        celda_mejor = f"${mejor_letra}${mejor_fila}"
    else:
        celda_mejor = None

    celdas_punt_costo = []
    for i in range(n_cols_total_max):
        col_punt = 4 + i
        letra_tot, fila_tot_ref = formula_total_refs[i]
        cell = ws.cell(row=fila, column=col_punt)
        if celda_mejor and i < n_prov:
            cell.value = f"=+{celda_mejor}/{letra_tot}{fila_tot_ref}*5"
        else:
            cell.value = 0
        cell.number_format = "0.00"
        cell.font = font_normal
        cell.alignment = Alignment(horizontal="center")
        cell.border = borde
        if i == 0:
            cell.fill = fill_rosado
        celdas_punt_costo.append((get_column_letter(col_punt), fila))
    fila += 1

    # Fila "Condicion de pago" -- EN BLANCO (manual)
    fila_pago = fila
    ws.cell(row=fila, column=2, value="Condici\u00f3n de pago").font = font_normal
    ws.cell(row=fila, column=2).border = borde
    cell_peso_pago = ws.cell(row=fila, column=3, value=peso_pago / 100)
    cell_peso_pago.number_format = "0%"
    cell_peso_pago.font = font_normal
    cell_peso_pago.alignment = Alignment(horizontal="center")
    cell_peso_pago.border = borde
    celdas_punt_pago = []
    for i in range(n_cols_total_max):
        col_punt = 4 + i
        cell = ws.cell(row=fila, column=col_punt)
        cell.border = borde
        cell.alignment = Alignment(horizontal="center")
        cell.font = font_normal
        cell.number_format = "0.00"
        if i == 1:
            cell.fill = fill_verde
        celdas_punt_pago.append((get_column_letter(col_punt), fila))
    fila += 1

    # Fila PROMEDIO = (peso_costo*punt_costo)+(peso_pago*punt_pago)
    ws.cell(row=fila, column=2, value="PROMEDIO").font = font_bold
    ws.cell(row=fila, column=2).border = borde
    cell_peso_total = ws.cell(row=fila, column=3, value=(peso_costo + peso_pago) / 100)
    cell_peso_total.number_format = "0%"
    cell_peso_total.font = font_bold
    cell_peso_total.alignment = Alignment(horizontal="center")
    cell_peso_total.border = borde
    for i in range(n_cols_total_max):
        col_punt = 4 + i
        letra_c, fila_c = celdas_punt_costo[i]
        letra_p, fila_p = celdas_punt_pago[i]
        formula = f"=($C${fila_costos}*{letra_c}{fila_c})+($C${fila_pago}*{letra_p}{fila_p})"
        cell = ws.cell(row=fila, column=col_punt, value=formula)
        cell.number_format = "0.00"
        cell.font = font_bold
        cell.alignment = Alignment(horizontal="center")
        cell.border = borde
        if i == 0:
            cell.fill = fill_rosado
        elif i == 1:
            cell.fill = fill_verde
    fila += 1

    ws.freeze_panes = "C1"

    # â”€â”€ SEGUNDA HOJA: Condiciones Comerciales â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws2 = wb.create_sheet(title="Condiciones")
    ws2.column_dimensions["A"].width = 32
    for i in range(n_prov):
        col_prov = 2 + i
        ws2.column_dimensions[get_column_letter(col_prov)].width = 30

    # Header banda naranja
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + n_prov)
    cell_h = ws2.cell(row=1, column=1, value="Condiciones Comerciales por Proveedor")
    cell_h.fill = fill_header
    cell_h.font = font_header
    cell_h.alignment = Alignment(horizontal="center")
    for c in range(1, 2 + n_prov):
        ws2.cell(row=1, column=c).fill = fill_header

    # Sub-header con nombres de proveedor
    ws2.cell(row=2, column=1, value="Criterio").font = font_bold
    ws2.cell(row=2, column=1).fill = PatternFill(start_color="FBE5D6", end_color="FBE5D6", fill_type="solid")
    ws2.cell(row=2, column=1).border = borde
    for i in range(n_prov):
        nombre_p = gv(df.iloc[i], col_nombre, f"Proveedor {i+1}")
        cell = ws2.cell(row=2, column=2 + i, value=str(nombre_p))
        cell.font = font_bold
        cell.fill = PatternFill(start_color="FBE5D6", end_color="FBE5D6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = borde
    ws2.row_dimensions[2].height = 36

    # Filas de condiciones
    filas_cond = [
        ("Garant\u00eda", "garantia"),
        ("Tiempo de entrega", "tiempo_entrega"),
        ("Condiciones de pago", "condiciones_pago"),
        ("Precio sin IGV", "precio_sin_igv"),
        ("Precio con IGV", "precio_con_igv"),
    ]
    fill_par  = PatternFill(start_color="FEF9F5", end_color="FEF9F5", fill_type="solid")
    fill_impar = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for fila_c, (label, campo) in enumerate(filas_cond, 3):
        ws2.cell(row=fila_c, column=1, value=label).font = font_bold
        ws2.cell(row=fila_c, column=1).border = borde
        ws2.cell(row=fila_c, column=1).fill = fill_par if fila_c % 2 == 0 else fill_impar
        max_h = 1
        for i in range(n_prov):
            valor = gv(df.iloc[i], campo, "No especifica")
            cell = ws2.cell(row=fila_c, column=2 + i, value=str(valor))
            cell.font = Font(size=10, name="Calibri", color=GRIS_LABEL)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = borde
            cell.fill = fill_par if fila_c % 2 == 0 else fill_impar
            lineas = max(1, len(str(valor)) // 28 + 1)
            if lineas > max_h:
                max_h = lineas
        ws2.row_dimensions[fila_c].height = max(20, min(max_h * 15, 90))

    ws2.freeze_panes = "B3"
    ws2.page_setup.orientation = "landscape"
    ws2.page_setup.fitToWidth = 1

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


CYAN   = ["#FFFFFF","#FFFFFF","#E5DFD3","#4A90B8","#3A7CA5","#2E3A45"]
VERDE  = ["#2E5E3D","#3D7A4D","#5C9C6E","#8AB89A","#B5D4C0","#E6F4E9"]
LAYOUT = dict(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", font_color="#2E3A45")

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# MODULO 1: CVs
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
if st.session_state.modulo_activo == "cvs":
    titulo_modulo("M\u00f3dulo de an\u00e1lisis de CVs", "Carga &amp; clasificaci\u00f3n por lotes")
    uploaded_files = st.file_uploader("Sube los CVs en PDF", type=["pdf"], accept_multiple_files=True)
    if uploaded_files:
        banner_archivos(len(uploaded_files), "archivos cargados, listos para procesar")

    if uploaded_files and st.button("Procesar y clasificar candidatos", type="primary", use_container_width=True):
        resultados = []
        progreso_holder = st.empty(); pb = st.progress(0); stxt = st.empty()
        habilidades_lista = [h.strip().lower() for h in habilidades_req.split("\n") if h.strip()]
        for idx, file in enumerate(uploaded_files):
            with progreso_holder.container():
                progreso_label("Procesando", idx+1, len(uploaded_files))
            stxt.markdown(f'<div style="font-family:var(--nx-mono);font-size:11px;color:#9CA8B0;margin-top:4px;">{file.name}</div>', unsafe_allow_html=True)
            try:
                reader = PdfReader(file)
                texto  = "".join(p.extract_text() or "" for p in reader.pages)
            except Exception as e:
                st.error(f"Error leyendo {file.name}: {e}"); continue
            prompt = (
                f'Eres un reclutador experto. Analiza este CV para el puesto de "{puesto or "No especificado"}".\n'
                f'Responde EXCLUSIVAMENTE con un objeto JSON valido, sin texto adicional ni markdown.\n\n'
                f'Formato JSON requerido:\n'
                f'{{"Nombre":"","Correo":"","Telefono":"","Educacion_Maxima":"","Universidad":"","Carrera":"",'
                f'"Ultimo_Cargo":"","Ultima_Empresa":"","Experiencia_Anos":0,"Habilidades_Tecnicas":"",'
                f'"Habilidades_Blandas":"","Idiomas":"","Certificaciones":"","Calculo_Interno":"",'
                f'"Puntaje":0,"Nivel_Potencial":"","Justificacion":"","Cumple_Requisitos":false,'
                f'"Requisitos_Cumplidos":"","Requisitos_Faltantes":""}}\n\n'
                f'Instrucciones:\n'
                f'- Calcula el Puntaje del 1 al 10 ponderando: experiencia ({peso_exp}%), educacion ({peso_edu}%), '
                f'habilidades: {", ".join(habilidades_lista) or "no especificadas"} ({peso_hab}%), '
                f'idiomas: {idioma_req} ({peso_idi}%)\n'
                f'- Calculo_Interno: muestra aqui el desglose matematico completo del calculo del puntaje '
                f'(formula, porcentajes, suma). Este campo es solo para verificacion interna, no se muestra al usuario.\n'
                f'- Nivel_Potencial: "Alto" si Puntaje >= 7, "Medio" si >= 4, "Bajo" si < 4\n'
                f'- Cumple_Requisitos: true si experiencia >= {experiencia_min} y Puntaje >= 6\n'
                f'- Experiencia_Anos: solo numero entero\n'
                f'- Habilidades_Tecnicas: maximo 6, separadas por coma\n'
                f'- Si un dato no existe: "No especifica"\n'
                f'- Justificacion: maximo 2 oraciones cortas (menos de 200 caracteres) explicando '
                f'la fortaleza y debilidad principal del candidato. NO incluyas numeros, formulas, '
                f'porcentajes ni el valor del puntaje en este campo (eso ya va en Calculo_Interno).\n\n'
                f'CV:\n{texto[:4000]}'
            )
            try:
                msg   = client.messages.create(model="claude-haiku-4-5-20251001", max_tokens=1200,
                            messages=[{"role":"user","content":prompt}])
                datos = limpiar_json(msg.content[0].text)
                # Quitar el campo de calculo interno (solo se usaba para que Claude razone el puntaje)
                datos.pop("Calculo_Interno", None)
                # Seguridad adicional: si Justificacion aun contiene restos de formulas, cortarlos
                if "Justificacion" in datos and isinstance(datos["Justificacion"], str):
                    just = datos["Justificacion"]
                    for marcador in ["Puntaje calculado", "calculo:", "C\u00e1lculo:", "= ", "+ (", "\u00d7"]:
                        if marcador in just:
                            just = just.split(marcador)[0].strip()
                    datos["Justificacion"] = just
                datos["Archivo"] = file.name; resultados.append(datos)
            except Exception as e:
                st.error(f"Error procesando {file.name}: {e}")
            pb.progress((idx+1)/len(uploaded_files))
        if resultados:
            df = pd.DataFrame(resultados)
            df["Experiencia_Anos"] = pd.to_numeric(df["Experiencia_Anos"], errors="coerce").fillna(0).astype(int)
            df["Puntaje"]          = pd.to_numeric(df["Puntaje"],          errors="coerce").fillna(0)
            df = df.sort_values("Puntaje", ascending=False).reset_index(drop=True)
            df["Ranking"] = df.index + 1
            st.session_state.df_candidatos = df
            stxt.success(f"{len(resultados)} CVs procesados correctamente.")

            titulo_hist = f"{len(df)} CVs"
            if puesto.strip():
                titulo_hist += f" \u00b7 {puesto.strip()}"
            guardar_historial("cvs", titulo_hist, df_a_json_seguro(df))
            st.session_state.historial_item_activo = None

    if st.session_state.df_candidatos is not None:
        df = st.session_state.df_candidatos.copy()
        tab1, tab2, tab3, tab4 = st.tabs(["Ranking","Dashboards","Filtros","Exportar"])

        with tab1:
            st.markdown("## Ranking de Candidatos")
            c1,c2,c3,c4 = st.columns(4)
            with c1: st.markdown(f'<div class="metric-card"><p>Total CVs</p><h2>{len(df)}</h2></div>', unsafe_allow_html=True)
            with c2: st.markdown(f'<div class="metric-card" style="border-top-color:#3D7A4D;"><p>Alto Potencial</p><h2>{len(df[df["Nivel_Potencial"]=="Alto"])}</h2></div>', unsafe_allow_html=True)
            with c3: st.markdown(f'<div class="metric-card" style="border-top-color:#A8762E;"><p>Potencial Medio</p><h2>{len(df[df["Nivel_Potencial"]=="Medio"])}</h2></div>', unsafe_allow_html=True)
            with c4: st.markdown(f'<div class="metric-card" style="border-top-color:#4A90B8;"><p>Cumplen Requisitos</p><h2>{len(df[df["Cumple_Requisitos"]==True])}</h2></div>', unsafe_allow_html=True)
            st.markdown('<div class="nx-section-label"><span>Ranking de candidatos</span><div class="nx-rule"></div></div>', unsafe_allow_html=True)
            for _, row in df.head(20).iterrows():
                nivel  = row.get("Nivel_Potencial","Bajo")
                badge  = "badge-alto" if nivel=="Alto" else "badge-medio" if nivel=="Medio" else "badge-bajo"
                cumple = "Cumple requisitos" if row.get("Cumple_Requisitos") else "No cumple"
                track_c, arc_c = dial_colores_cv(nivel)
                dial = render_dial(row.get("Puntaje",0), track_c, arc_c)
                rank_num = f"{int(row['Ranking']):02d}"
                st.markdown(f"""<div class="nx-row">
                  <div class="nx-rank">{rank_num}</div>
                  {dial}
                  <div style="flex:1;">
                    <div style="display:flex;align-items:center;gap:10px;margin-bottom:5px;flex-wrap:wrap;">
                      <span class="nx-name">{esc(row.get('Nombre','N/A'))}</span>
                      <span class="{badge}">{esc(nivel)}</span>
                      <span style="font-family:var(--nx-mono);font-size:10px;color:#9CA8B0;text-transform:uppercase;letter-spacing:1px;">{cumple}</span>
                    </div>
                    <div class="nx-meta">{esc(row.get('Ultimo_Cargo','N/A'))} &middot; {esc(row.get('Ultima_Empresa','N/A'))} &nbsp;/&nbsp; {row.get('Experiencia_Anos',0)} a\u00f1os exp &nbsp;/&nbsp; {esc(row.get('Educacion_Maxima','N/A'))}</div>
                    <div class="nx-desc">{esc(row.get('Justificacion',''))}</div>
                  </div>
                  <div class="nx-chev"><i class="ti ti-chevron-right" style="font-size:18px;"></i></div>
                </div>""", unsafe_allow_html=True)

        with tab2:
            st.markdown("## Dashboards y M\u00e9tricas")
            col_a,col_b = st.columns(2)
            with col_a:
                cnt = df["Nivel_Potencial"].value_counts().reset_index(); cnt.columns=["Nivel","Cantidad"]
                fig1 = px.pie(cnt,values="Cantidad",names="Nivel",title="Distribuci\u00f3n por Potencial",hole=0.45,
                              color="Nivel",color_discrete_map={"Alto":"#4ade80","Medio":"#fbbf24","Bajo":"#f87171"})
                fig1.update_layout(**LAYOUT); st.plotly_chart(fig1,use_container_width=True)
            with col_b:
                edu = df["Educacion_Maxima"].value_counts().reset_index(); edu.columns=["Educaci\u00f3n","Cantidad"]
                fig2 = px.bar(edu,x="Cantidad",y="Educaci\u00f3n",orientation="h",title="Nivel Educativo",
                              color="Cantidad",color_continuous_scale=CYAN)
                fig2.update_layout(**LAYOUT); st.plotly_chart(fig2,use_container_width=True)
            col_c,col_d = st.columns(2)
            with col_c:
                all_hab = []
                for h in df["Habilidades_Tecnicas"].dropna():
                    all_hab.extend([x.strip() for x in str(h).split(",") if x.strip() and x.strip()!="No especifica"])
                if all_hab:
                    df_h = pd.DataFrame(Counter(all_hab).most_common(12),columns=["Habilidad","Frecuencia"])
                    fig3 = px.bar(df_h,x="Frecuencia",y="Habilidad",orientation="h",title="Habilidades Tecnicas",
                                  color="Frecuencia",color_continuous_scale=CYAN)
                    fig3.update_layout(**LAYOUT); st.plotly_chart(fig3,use_container_width=True)
            with col_d:
                fig4 = px.scatter(df,x="Experiencia_Anos",y="Puntaje",color="Nivel_Potencial",
                                  hover_data=["Nombre","Ultimo_Cargo"],title="Experiencia vs Puntaje",
                                  color_discrete_map={"Alto":"#4ade80","Medio":"#fbbf24","Bajo":"#f87171"})
                fig4.update_layout(**LAYOUT); st.plotly_chart(fig4,use_container_width=True)
            k1,k2,k3,k4 = st.columns(4)
            with k1: st.metric("Experiencia Promedio",f"{df['Experiencia_Anos'].mean():.1f} a\u00f1os")
            with k2: st.metric("Puntaje Promedio",    f"{df['Puntaje'].mean():.1f}/10")
            with k3: st.metric("Cumple Requisitos",   f"{len(df[df['Cumple_Requisitos']==True])/len(df)*100:.0f}%")
            with k4: st.metric("Alto Potencial",      f"{len(df[df['Nivel_Potencial']=='Alto'])/len(df)*100:.0f}%")

        with tab3:
            st.markdown("## Filtros en Tiempo Real")
            f1,f2,f3,f4 = st.columns(4)
            with f1: fp  = st.multiselect("Potencial",["Alto","Medio","Bajo"],default=["Alto","Medio","Bajo"])
            with f2: fe  = st.slider("Exp. m\u00ednima (a\u00f1os)",0,20,0)
            with f3: fpu = st.slider("Puntaje m\u00ednimo",0.0,10.0,0.0,0.5)
            with f4: fc  = st.selectbox("Cumple req.",["Todos","Solo los que cumplen","Solo los que no cumplen"])
            fn = st.text_input("Buscar nombre, cargo o habilidad")
            df_f = df[df["Nivel_Potencial"].isin(fp)]
            df_f = df_f[df_f["Experiencia_Anos"]>=fe]
            df_f = df_f[df_f["Puntaje"]>=fpu]
            if fc=="Solo los que cumplen":     df_f=df_f[df_f["Cumple_Requisitos"]==True]
            elif fc=="Solo los que no cumplen": df_f=df_f[df_f["Cumple_Requisitos"]==False]
            if fn:
                mask=(df_f["Nombre"].str.contains(fn,case=False,na=False)|
                      df_f["Ultimo_Cargo"].str.contains(fn,case=False,na=False)|
                      df_f["Habilidades_Tecnicas"].str.contains(fn,case=False,na=False))
                df_f=df_f[mask]
            st.markdown(f"**{len(df_f)} candidatos encontrados**")
            cols_show=["Ranking","Nombre","Puntaje","Nivel_Potencial","Ultimo_Cargo",
                       "Experiencia_Anos","Educacion_Maxima","Habilidades_Tecnicas","Idiomas","Correo","Cumple_Requisitos"]
            st.dataframe(df_f[[c for c in cols_show if c in df_f.columns]],use_container_width=True,height=500)

        with tab4:
            st.markdown("## Exportar Resultados")
            op = st.radio("\u00bfQu\u00e9 candidatos exportar?",["Todos","Solo Alto Potencial","Solo los que cumplen","Top 10"])
            df_exp = df.copy()
            if op=="Solo Alto Potencial":    df_exp=df[df["Nivel_Potencial"]=="Alto"]
            elif op=="Solo los que cumplen": df_exp=df[df["Cumple_Requisitos"]==True]
            elif op=="Top 10":               df_exp=df.head(10)
            st.info(f"Se exportaran **{len(df_exp)} candidatos**")
            st.download_button("Descargar Excel",data=exportar_excel_cvs(df_exp),
                file_name="Nexora_CVs.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# MODULO 2: PROVEEDORES
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
elif st.session_state.modulo_activo == "proveedores":
    titulo_modulo("M\u00f3dulo de an\u00e1lisis de proveedores", "B\u00fasqueda, documentos &amp; comparaci\u00f3n")
    tab_buscar,tab_subir,tab_comparar,tab_dash,tab_export = st.tabs([
        "Buscar en Internet","Analizar Documentos","Comparar Proveedores","Dashboards","Exportar"
    ])

    with tab_buscar:
        st.markdown("### B\u00fasqueda de Proveedores en Internet")
        st.markdown("<p style='color:#3A7CA5;'>Describe con todo el detalle que necesites.</p>", unsafe_allow_html=True)
        col_q1,col_q2 = st.columns([3,1])
        with col_q1:
            query_usuario = st.text_area("Describe qu\u00e9 proveedor necesitas",
                placeholder="Ej: Necesito proveedores de servicios de nomina para una empresa de 200 empleados en Peru.",height=120)
        with col_q2:
            num_resultados   = st.selectbox("Cantidad de proveedores",[5,8,10,15],index=1)
            incluir_precios  = st.checkbox("Incluir rango de precios",value=True)
            incluir_contacto = st.checkbox("Incluir contacto/web",value=True)

        if st.button("Buscar Proveedores",type="primary",use_container_width=True):
            if not query_usuario.strip():
                st.warning("Por favor describe qu\u00e9 proveedor necesitas.")
            else:
                contexto=(f"Pais: {pais_busqueda or 'no especificado'} | Rubro: {rubro_busqueda or 'no especificado'} | "
                          f"Presupuesto: {presupuesto_ref} | Cobertura: {cobertura_req}")

                TAMANO_LOTE = 4
                lotes = []
                restante = num_resultados
                while restante > 0:
                    lotes.append(min(TAMANO_LOTE, restante))
                    restante -= TAMANO_LOTE

                proveedores_detallados=[]
                resumen_mercado=""
                nombres_ya_encontrados=[]
                pb = st.progress(0)
                status = st.empty()

                for idx_lote, cantidad_lote in enumerate(lotes):
                    status.markdown(f"Buscando lote {idx_lote+1}/{len(lotes)} ({cantidad_lote} proveedores)...")

                    exclusion = ""
                    if nombres_ya_encontrados:
                        exclusion = (f"\nNO incluyas estas empresas que ya fueron encontradas: "
                                      f"{', '.join(nombres_ya_encontrados)}.")

                    prompt_busqueda=(
                        f"Busca en internet empresas reales y verificadas que ofrezcan: {query_usuario}\n"
                        f"Contexto: {contexto}{exclusion}\n\n"
                        f"Encuentra exactamente {cantidad_lote} empresas reales (con sitio web verificable) "
                        f"y para CADA UNA devuelve la informacion completa que se pide abajo, "
                        f"basandote en lo que encuentres en internet sobre cada empresa.\n\n"
                        f"Responde SOLO con JSON valido, sin texto adicional ni markdown:\n"
                        f'{{"proveedores":[{{"nombre":"Nombre real","descripcion":"descripcion completa de la empresa",'
                        f'"sitio_web":"https://...","pais_sede":"pais","cobertura":"Local o Nacional o Regional o Internacional",'
                        f'"anos_experiencia":"numero o rango","certificaciones":"lista o No especifica",'
                        f'"rango_precio":"rango estimado o No publico","contacto":"email o telefono real",'
                        f'"fortalezas":"f1, f2, f3","clientes_referencia":"clientes conocidos o No publico",'
                        f'"puntaje_recomendacion":8,"nivel_recomendacion":"Muy recomendado",'
                        f'"razon_recomendacion":"razon especifica para este caso"}}],'
                        f'"resumen_mercado":"resumen ejecutivo breve (2-3 oraciones) del mercado de proveedores encontrado"}}'
                    )

                    try:
                        msg=client.messages.create(model="claude-haiku-4-5-20251001",max_tokens=4096,
                            tools=[{"type":"web_search_20250305","name":"web_search"}],
                            messages=[{"role":"user","content":prompt_busqueda}])
                        texto_resp="".join(b.text for b in msg.content if hasattr(b,"text"))
                        datos=limpiar_json(texto_resp)
                        nuevos = datos.get("proveedores",[])
                        proveedores_detallados.extend(nuevos)
                        nombres_ya_encontrados.extend([p.get("nombre","") for p in nuevos if p.get("nombre")])
                        if not resumen_mercado:
                            resumen_mercado = datos.get("resumen_mercado","")
                    except Exception as e:
                        st.warning(f"Lote {idx_lote+1} fallo: {e}")

                    pb.progress((idx_lote+1)/len(lotes))

                status.empty()
                pb.empty()

                if proveedores_detallados:
                    if resumen_mercado:
                        st.info(f"An\u00e1lisis del mercado: {resumen_mercado}")
                    st.session_state.proveedores_web=proveedores_detallados
                    st.success(f"Busqueda completada: {len(proveedores_detallados)} proveedores analizados.")

                    resumen_query = query_usuario.strip()[:50]
                    if len(query_usuario.strip()) > 50:
                        resumen_query += "\u2026"
                    titulo_hist = f"{len(proveedores_detallados)} proveedores \u00b7 {resumen_query}"
                    guardar_historial("proveedores", titulo_hist, {
                        "proveedores_web": proveedores_detallados,
                        "df_proveedores": None,
                    })
                    st.session_state.historial_item_activo = None
                else:
                    st.error("No se pudo completar la busqueda. Intenta de nuevo o reduce la cantidad de proveedores.")

        if st.session_state.proveedores_web:
            provs=st.session_state.proveedores_web
            provs_sorted=sorted(provs,key=lambda x:safe_float(x.get("puntaje_recomendacion",0)),reverse=True)
            st.markdown(f'<div class="nx-section-label"><span>{len(provs_sorted)} proveedores encontrados</span><div class="nx-rule"></div></div>', unsafe_allow_html=True)
            for i,prov in enumerate(provs_sorted,1):
                puntaje_num=safe_float(prov.get("puntaje_recomendacion",0))
                nivel_texto=prov.get("nivel_recomendacion","Recomendado")
                badge_c="badge-prov-a" if puntaje_num>=8 else "badge-prov-b" if puntaje_num>=6 else "badge-prov-c"
                track_c, arc_c = dial_colores_prov(puntaje_num)
                dial = render_dial(puntaje_num, track_c, arc_c)
                rank_num = f"{i:02d}"
                st.markdown(f"""<div class="nx-row">
                  <div class="nx-rank">{rank_num}</div>
                  {dial}
                  <div style="flex:1;">
                    <div style="display:flex;align-items:center;gap:10px;margin-bottom:5px;flex-wrap:wrap;">
                      <span class="nx-name">{esc(prov.get('nombre','N/A'))}</span>
                      <span class="{badge_c}">{esc(nivel_texto)}</span>
                    </div>
                    <div class="nx-meta">{esc(prov.get('pais_sede','N/A'))} &middot; Cobertura {esc(prov.get('cobertura','N/A'))} &nbsp;/&nbsp; {esc(prov.get('anos_experiencia','N/A'))} a\u00f1os exp &nbsp;/&nbsp; {esc(prov.get('certificaciones','N/A'))}</div>
                    <div class="nx-desc">{esc(prov.get('descripcion',''))}</div>
                    <div style="margin-top:8px;font-family:var(--nx-mono);font-size:11px;color:#9CA8B0;display:flex;flex-wrap:wrap;gap:14px;">
                      <span>Precio: {esc(prov.get('rango_precio','N/A'))}</span>
                      <span>Web: <a href="{safe_url(prov.get('sitio_web','#'))}" target="_blank" style="color:#3A7CA5;">{esc(prov.get('sitio_web','N/A'))}</a></span>
                      <span>Contacto: {esc(prov.get('contacto','N/A'))}</span>
                    </div>
                    <div style="margin-top:6px;font-size:12px;color:#3D7A4D;">{esc(prov.get('fortalezas',''))}</div>
                    <div style="margin-top:4px;font-size:11px;color:#4A90B8;font-style:italic;">{esc(prov.get('razon_recomendacion',''))}</div>
                  </div>
                  <div class="nx-chev"><i class="ti ti-chevron-right" style="font-size:18px;"></i></div>
                </div>""", unsafe_allow_html=True)
            if st.button("Agregar al comparador",use_container_width=True):
                df_web=pd.DataFrame(provs_sorted)
                if st.session_state.df_proveedores is None: st.session_state.df_proveedores=df_web
                else:
                    st.session_state.df_proveedores=pd.concat(
                        [st.session_state.df_proveedores,df_web],ignore_index=True
                    ).drop_duplicates(subset=["nombre"])
                st.success("Proveedores agregados al comparador.")

    with tab_subir:
        st.markdown("### Analizar Documentos de Proveedores")
        st.markdown("<p style='color:#3A7CA5;'>Sube propuestas, RFPs o fichas tecnicas en PDF.</p>", unsafe_allow_html=True)
        docs=st.file_uploader("Sube documentos PDF",type=["pdf"],accept_multiple_files=True,key="docs_prov")
        if docs: banner_archivos(len(docs), "documentos cargados, listos para analizar")
        if docs and st.button("Analizar documentos",type="primary",use_container_width=True):
            resultados_prov=[]; progreso_holder2=st.empty(); pb2=st.progress(0); stx2=st.empty()
            cert_lista=[c.strip() for c in cert_requeridas.split("\n") if c.strip()]
            for idx,doc in enumerate(docs):
                with progreso_holder2.container():
                    progreso_label("Analizando", idx+1, len(docs))
                stx2.markdown(f'<div style="font-family:var(--nx-mono);font-size:11px;color:#9CA8B0;margin-top:4px;">{doc.name}</div>', unsafe_allow_html=True)
                try:
                    reader=PdfReader(doc); texto="".join(p.extract_text() or "" for p in reader.pages)
                except Exception as e:
                    st.error(f"Error leyendo {doc.name}: {e}"); continue
                prompt_doc=(f"Eres un experto en procurement. Analiza este documento de propuesta de proveedor.\n"
                            f"Responde EXCLUSIVAMENTE con JSON valido, sin texto adicional ni markdown.\n\n"
                            f"Contexto: pais={pais_busqueda or 'no especificado'}, presupuesto={presupuesto_ref}, cobertura={cobertura_req}\n"
                            f"Certificaciones requeridas: {', '.join(cert_lista) or 'ninguna'}\n"
                            f"Pesos: Precio {ppeso_precio}%, Cert {ppeso_cert}%, Rep {ppeso_rep}%, Cob {ppeso_cob}%\n\n"
                            f'{{"nombre":"","descripcion":"","sitio_web":"","pais_sede":"","cobertura":"",'
                            f'"anos_experiencia":"","certificaciones":"","productos_servicios":"",'
                            f'"rango_precio":"","precio_sin_igv":"","precio_con_igv":"",'
                            f'"condiciones_pago":"","tiempo_entrega":"","garantia":"",'
                            f'"condiciones_comerciales":"","clientes_referencia":"","fortalezas":"","debilidades":"",'
                            f'"item_descripcion":"","item_cantidad":1,"moneda":"","costo_unitario_num":0,"costo_total_num":0,'
                            f'"puntaje_precio":0,"puntaje_certificaciones":0,"puntaje_reputacion":0,"puntaje_cobertura":0,'
                            f'"puntaje_recomendacion":0,"nivel_recomendacion":"","justificacion":"",'
                            f'"cumple_certificaciones":false,"certificaciones_faltantes":""}}\n\n'
                            f'Presta ESPECIAL ATENCION a extraer con precision estos puntos si aparecen en el documento:\n'
                            f'- garantia: duracion y cobertura de la garantia ofrecida (ej: "12 meses contra defectos de fabricacion")\n'
                            f'- tiempo_entrega: plazo de entrega del producto/servicio (ej: "15 dias habiles")\n'
                            f'- precio_sin_igv: precio del producto/servicio SIN IGV/IVA, incluye moneda (ej: "S/ 8,500.00")\n'
                            f'- precio_con_igv: precio del producto/servicio CON IGV/IVA, incluye moneda (ej: "S/ 10,030.00")\n'
                            f'- condiciones_pago: forma y plazos de pago (ej: "50% adelanto, 50% contra entrega")\n'
                            f'- item_descripcion: nombre/descripcion breve (max 80 caracteres) del producto o servicio principal cotizado\n'
                            f'- item_cantidad: cantidad cotizada del item principal, solo numero entero (default 1 si no se especifica)\n'
                            f'- moneda: codigo de moneda detectado (PEN, USD, etc). Usa PEN si ves S/. o Soles, USD si ves $ o Dolares\n'
                            f'- costo_unitario_num: el precio_sin_igv como NUMERO puro (sin simbolos, sin comas de miles, usa punto decimal). '
                            f'Ejemplo: si precio_sin_igv es "S/.41,000.00", costo_unitario_num debe ser 41000.00\n'
                            f'- costo_total_num: costo_unitario_num multiplicado por item_cantidad (si cantidad es 1, son iguales)\n\n'
                            f'- nivel_recomendacion: "Muy recomendado">=8, "Recomendado">=6, "Opcion viable">=4, "No recomendado"<4\n'
                            f"- Si un dato no aparece en el documento, escribe: No especifica (para campos numericos usa 0)\n\n"
                            f"Documento:\n{preparar_texto_documento(texto)}")
                try:
                    msg=client.messages.create(model="claude-haiku-4-5-20251001",max_tokens=1500,
                            messages=[{"role":"user","content":prompt_doc}])
                    datos=limpiar_json(msg.content[0].text)
                    datos["Archivo"]=doc.name; datos["Fuente"]="Documento"; resultados_prov.append(datos)
                except Exception as e:
                    st.error(f"Error procesando {doc.name}: {e}")
                pb2.progress((idx+1)/len(docs))
            if resultados_prov:
                df_new=pd.DataFrame(resultados_prov)
                if st.session_state.df_proveedores is None: st.session_state.df_proveedores=df_new
                else: st.session_state.df_proveedores=pd.concat([st.session_state.df_proveedores,df_new],ignore_index=True)
                st.session_state.df_propuestas=df_new
                stx2.success(f"{len(resultados_prov)} documentos analizados y agregados al comparador.")

                col_nombre_hist = "nombre" if "nombre" in df_new.columns else None
                if col_nombre_hist:
                    nombres_prop = ", ".join(df_new[col_nombre_hist].dropna().astype(str).head(3).tolist())
                else:
                    nombres_prop = ", ".join(df_new["Archivo"].dropna().astype(str).head(3).tolist())
                titulo_hist = f"{len(resultados_prov)} propuestas \u00b7 {nombres_prop}"
                guardar_historial("propuestas", titulo_hist, resultados_prov)
                st.session_state.historial_item_activo = None

        if st.session_state.get("df_propuestas") is not None and len(st.session_state.df_propuestas)>0:
            st.divider()
            st.markdown("#### Condiciones Comerciales Extraidas")
            st.markdown("<p style='color:#3A7CA5;font-size:13px;'>Garant\u00eda, tiempo de entrega, precios con/sin IGV y condiciones de pago detectados en cada propuesta.</p>", unsafe_allow_html=True)

            df_prop = st.session_state.df_propuestas.copy()
            col_nombre = "nombre" if "nombre" in df_prop.columns else "nombre_empresa"
            cols_resumen = [col_nombre,"Archivo","garantia","tiempo_entrega","precio_sin_igv","precio_con_igv","condiciones_pago"]
            cols_resumen_ok = [c for c in cols_resumen if c in df_prop.columns]
            rename_resumen = {col_nombre:"Proveedor","Archivo":"Archivo","garantia":"Garant\u00eda",
                "tiempo_entrega":"Tiempo de Entrega","precio_sin_igv":"Precio sin IGV",
                "precio_con_igv":"Precio con IGV","condiciones_pago":"Condiciones de Pago"}
            st.dataframe(df_prop[cols_resumen_ok].rename(columns=rename_resumen), use_container_width=True, height=250)

            cols_excel = [c for c in [col_nombre,"Archivo","garantia","tiempo_entrega","precio_sin_igv","precio_con_igv","condiciones_pago","rango_precio","condiciones_comerciales"] if c in df_prop.columns]
            df_excel_prop = df_prop[cols_excel].copy()
            if col_nombre != "nombre":
                df_excel_prop = df_excel_prop.rename(columns={col_nombre:"nombre"})

            with st.expander("Datos opcionales para el encabezado de la Matriz Econ\u00f3mica"):
                desc_matriz = st.text_input("Descripci\u00f3n de compra", key="desc_matriz_input",
                    placeholder="Ej: Reemplazo l\u00ednea de gases del laboratorio - PR209244").strip()
                comprador_matriz = st.text_input("Comprador", key="comprador_matriz_input",
                    placeholder="Nombre del comprador responsable").strip()

            col_dl1, col_dl2 = st.columns(2)
            with col_dl1:
                st.download_button("Descargar Excel de Condiciones Comerciales",
                    data=exportar_excel_propuestas(df_excel_prop),
                    file_name="Nexora_Condiciones_Comerciales.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
            with col_dl2:
                df_matriz = df_prop.copy()
                if col_nombre != "nombre":
                    df_matriz = df_matriz.rename(columns={col_nombre:"nombre"})
                st.download_button("Descargar Matriz Econ\u00f3mica",
                    data=exportar_excel_matriz_economica(df_matriz, peso_matriz_costo, peso_matriz_pago,
                        username, descripcion_compra=desc_matriz, comprador=comprador_matriz),
                    file_name="Nexora_Matriz_Economica.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)

    with tab_comparar:
        st.markdown("### Comparaci\u00f3n de Proveedores")
        if st.session_state.df_proveedores is None or len(st.session_state.df_proveedores)==0:
            st.info("Primero busca proveedores en internet o analiza documentos.")
        else:
            df_p=st.session_state.df_proveedores.copy()
            if "nombre" not in df_p.columns and "nombre_empresa" in df_p.columns:
                df_p["nombre"]=df_p["nombre_empresa"]
            for col in ["puntaje_recomendacion","puntaje_precio","puntaje_certificaciones","puntaje_reputacion","puntaje_cobertura"]:
                if col in df_p.columns:
                    df_p[col]=pd.to_numeric(df_p[col],errors="coerce").fillna(0)
            todos_nombres=df_p["nombre"].dropna().tolist()
            seleccionados=st.multiselect("Selecciona proveedores a comparar",
                todos_nombres,default=todos_nombres[:min(5,len(todos_nombres))])
            if seleccionados:
                df_sel=df_p[df_p["nombre"].isin(seleccionados)].copy()
                if "puntaje_recomendacion" in df_sel.columns:
                    df_sel=df_sel.sort_values("puntaje_recomendacion",ascending=False).reset_index(drop=True)
                    df_sel["Ranking"]=df_sel.index+1
                st.markdown('<div class="nx-section-label"><span>Ranking general</span><div class="nx-rule"></div></div>', unsafe_allow_html=True)
                for _,row in df_sel.iterrows():
                    puntaje=row.get("puntaje_recomendacion",0)
                    nivel=row.get("nivel_recomendacion","Recomendado")
                    badge_c="badge-prov-a" if puntaje>=8 else "badge-prov-b" if puntaje>=6 else "badge-prov-c"
                    desc=str(row.get("descripcion",row.get("productos_servicios","")))[:160]
                    just=row.get("justificacion",row.get("razon_recomendacion",""))
                    tiene_puntajes=any(row.get(c,0) not in [0,"-",None,""] for c in ["puntaje_precio","puntaje_certificaciones","puntaje_reputacion","puntaje_cobertura"])
                    if tiene_puntajes:
                        sub_items = [
                            ("Precio", row.get("puntaje_precio","-")),
                            ("Certif.", row.get("puntaje_certificaciones","-")),
                            ("Reput.", row.get("puntaje_reputacion","-")),
                            ("Cobert.", row.get("puntaje_cobertura","-")),
                        ]
                    else:
                        sub_items = [
                            ("Cobertura", row.get("cobertura","N/A")),
                            ("Experiencia", f'{row.get("anos_experiencia","N/A")} a\u00f1os'),
                            ("Precio", row.get("rango_precio","N/A")),
                            ("Certif.", str(row.get("certificaciones","N/A"))[:24]),
                        ]
                    sub_html = "".join(
                        f'<span style="margin-right:16px;"><span style="color:#B4B2A9;">{esc(lbl)}</span> {esc(val)}</span>'
                        for lbl, val in sub_items
                    )
                    track_c, arc_c = dial_colores_prov(puntaje)
                    dial = render_dial(puntaje, track_c, arc_c)
                    rank_num = f"{int(row.get('Ranking',0)):02d}"
                    st.markdown(f"""<div class="nx-row">
                      <div class="nx-rank">{rank_num}</div>
                      {dial}
                      <div style="flex:1;">
                        <div style="display:flex;align-items:center;gap:10px;margin-bottom:5px;flex-wrap:wrap;">
                          <span class="nx-name">{esc(row.get('nombre','N/A'))}</span>
                          <span class="{badge_c}">{esc(nivel)}</span>
                        </div>
                        <div class="nx-desc" style="margin-bottom:8px;">{esc(desc)}</div>
                        <div style="font-family:var(--nx-mono);font-size:11px;color:#9CA8B0;">{sub_html}</div>
                        <div style="margin-top:6px;font-size:11px;color:#4A90B8;font-style:italic;">{esc(just)}</div>
                      </div>
                      <div class="nx-chev"><i class="ti ti-chevron-right" style="font-size:18px;"></i></div>
                    </div>""", unsafe_allow_html=True)
                criterios_all=["puntaje_precio","puntaje_certificaciones","puntaje_reputacion","puntaje_cobertura"]
                criterios_ok=[c for c in criterios_all if c in df_sel.columns and df_sel[c].sum()>0]
                if len(df_sel)>1 and criterios_ok:
                    st.markdown("#### Radar de Comparaci\u00f3n")
                    labels={"puntaje_precio":"Precio","puntaje_certificaciones":"Certif.","puntaje_reputacion":"Reput.","puntaje_cobertura":"Cobert."}
                    cats=[labels.get(c,c) for c in criterios_ok]
                    colores=["#3A7CA5","#34d399","#fbbf24","#f87171","#60a5fa"]
                    fig_r=go.Figure()
                    for i,(_, row) in enumerate(df_sel.iterrows()):
                        vals=[float(row.get(c,0)) for c in criterios_ok]
                        fig_r.add_trace(go.Scatterpolar(r=vals+[vals[0]],theta=cats+[cats[0]],fill='toself',
                            name=str(row.get("nombre","Proveedor")),line_color=colores[i%len(colores)],
                            fillcolor=colores[i%len(colores)],opacity=0.3))
                    fig_r.update_layout(polar=dict(radialaxis=dict(visible=True,range=[0,10],color="#3A7CA5"),
                        bgcolor="#FFFFFF",angularaxis=dict(color="#3A7CA5")),
                        paper_bgcolor="rgba(0,0,0,0)",font_color="#2E3A45",
                        legend=dict(bgcolor="#FFFFFF",bordercolor="#4A90B8",borderwidth=1),height=450)
                    st.plotly_chart(fig_r,use_container_width=True)
                st.markdown("#### Tabla Comparativa Detallada")
                cols_tabla=["nombre","cobertura","anos_experiencia","certificaciones","rango_precio","contacto","sitio_web","fortalezas","puntaje_recomendacion","nivel_recomendacion"]
                cols_ok=[c for c in cols_tabla if c in df_sel.columns]
                rename_map={"nombre":"Nombre","cobertura":"Cobertura","anos_experiencia":"A\u00f1os de Experiencia",
                            "certificaciones":"Certificaciones","rango_precio":"Rango de Precio",
                            "contacto":"Contacto","sitio_web":"Sitio Web","fortalezas":"Fortalezas",
                            "puntaje_recomendacion":"Puntaje","nivel_recomendacion":"Nivel de Recomendaci\u00f3n"}
                st.dataframe(df_sel[cols_ok].rename(columns=rename_map),use_container_width=True,height=300)
            if st.button("Limpiar comparador",use_container_width=True):
                st.session_state.df_proveedores=None; st.session_state.proveedores_web=[]; st.rerun()

    with tab_dash:
        st.markdown("### Dashboards de Proveedores")
        if st.session_state.df_proveedores is None or len(st.session_state.df_proveedores)==0:
            st.info("Agrega proveedores desde las pestanas anteriores para ver los dashboards.")
        else:
            df_p=st.session_state.df_proveedores.copy()
            if "nombre" not in df_p.columns and "nombre_empresa" in df_p.columns: df_p["nombre"]=df_p["nombre_empresa"]
            for col in ["puntaje_recomendacion","puntaje_precio","puntaje_certificaciones","puntaje_reputacion","puntaje_cobertura"]:
                if col in df_p.columns: df_p[col]=pd.to_numeric(df_p[col],errors="coerce").fillna(0)
            col1,col2=st.columns(2)
            with col1:
                if "cobertura" in df_p.columns:
                    cob=df_p["cobertura"].value_counts().reset_index(); cob.columns=["Cobertura","Cantidad"]
                    fig_c=px.pie(cob,values="Cantidad",names="Cobertura",title="Distribuci\u00f3n por Cobertura",hole=0.4,color_discrete_sequence=VERDE)
                    fig_c.update_layout(**LAYOUT); st.plotly_chart(fig_c,use_container_width=True)
            with col2:
                if "nivel_recomendacion" in df_p.columns:
                    niv=df_p["nivel_recomendacion"].value_counts().reset_index(); niv.columns=["Nivel","Cantidad"]
                    fig_n=px.bar(niv,x="Nivel",y="Cantidad",title="Proveedores por Nivel",color="Cantidad",color_continuous_scale=VERDE)
                    fig_n.update_layout(**LAYOUT); st.plotly_chart(fig_n,use_container_width=True)
            if "puntaje_recomendacion" in df_p.columns and "nombre" in df_p.columns:
                df_rank=df_p[["nombre","puntaje_recomendacion"]].copy()
                df_rank=df_rank.sort_values("puntaje_recomendacion",ascending=True).tail(10)
                fig_rank=px.bar(df_rank,x="puntaje_recomendacion",y="nombre",orientation="h",
                                title="Top Proveedores por Puntaje",color="puntaje_recomendacion",color_continuous_scale=CYAN)
                fig_rank.update_layout(**LAYOUT); st.plotly_chart(fig_rank,use_container_width=True)
            k1,k2,k3=st.columns(3)
            with k1: st.metric("Total Proveedores",len(df_p))
            with k2:
                if "puntaje_recomendacion" in df_p.columns:
                    st.metric("Puntaje Promedio",f"{df_p['puntaje_recomendacion'].mean():.1f}/10")
            with k3:
                if "nivel_recomendacion" in df_p.columns:
                    st.metric("Muy Recomendados",len(df_p[df_p["nivel_recomendacion"].str.contains("Muy",na=False)]))

    with tab_export:
        st.markdown("### Exportar An\u00e1lisis de Proveedores")
        if st.session_state.df_proveedores is None or len(st.session_state.df_proveedores)==0:
            st.info("No hay proveedores para exportar aun.")
        else:
            df_p=st.session_state.df_proveedores.copy()
            if "nombre" not in df_p.columns and "nombre_empresa" in df_p.columns: df_p["nombre"]=df_p["nombre_empresa"]
            if "puntaje_recomendacion" in df_p.columns:
                df_p["puntaje_recomendacion"]=pd.to_numeric(df_p["puntaje_recomendacion"],errors="coerce").fillna(0)
            op_p=st.radio("\u00bfQu\u00e9 exportar?",["Todos los proveedores","Solo Muy Recomendados","Solo Recomendados","Top 5"])
            df_exp_p=df_p.copy()
            if "nivel_recomendacion" in df_p.columns:
                if op_p=="Solo Muy Recomendados": df_exp_p=df_p[df_p["nivel_recomendacion"].str.contains("Muy",na=False)]
                elif op_p=="Solo Recomendados": df_exp_p=df_p[df_p["nivel_recomendacion"].isin(["Muy recomendado","Recomendado","Muy Recomendado"])]
            if op_p=="Top 5":
                df_exp_p=df_p.sort_values("puntaje_recomendacion",ascending=False).head(5) if "puntaje_recomendacion" in df_p.columns else df_p.head(5)
            st.info(f"Se exportaran **{len(df_exp_p)} proveedores**")
            st.download_button("Descargar Excel de Proveedores",data=exportar_excel_proveedores(df_exp_p),
                file_name="Nexora_Proveedores.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)

# ── MODULO 3: ANÁLISIS DE CONTRATOS ─────────────────────────────────────────
elif st.session_state.modulo_activo == "contratos":
    titulo_modulo("Análisis de Contratos", "Extrae fechas, penalidades y riesgos con IA")
    client = anthropic.Anthropic(api_key=st.secrets["CLAUDE_API_KEY"])

    if "contrato_resultado" not in st.session_state:
        st.session_state.contrato_resultado = None

    tab_subir, tab_resultado = st.tabs(["📄 Subir contrato", "📊 Resultado"])

    with tab_subir:
        archivo = st.file_uploader("Sube el contrato (PDF o DOCX)", type=["pdf", "docx"],
                                   help="El archivo se analiza localmente — no se almacena")
        st.info("Claude leerá el contrato completo y extraerá fechas, montos, penalidades, obligaciones y nivel de riesgo.")

        if archivo and st.button("Analizar contrato", type="primary", use_container_width=True):
            with st.spinner("Analizando contrato..."):
                try:
                    if archivo.name.lower().endswith(".pdf"):
                        reader = PdfReader(archivo)
                        texto_contrato = "\n".join(p.extract_text() or "" for p in reader.pages)
                    else:
                        texto_contrato = archivo.read().decode("utf-8", errors="ignore")

                    prompt_contrato = (
                        f"Eres un abogado corporativo experto en análisis de contratos comerciales.\n"
                        f"Analiza el siguiente contrato y responde EXCLUSIVAMENTE con JSON válido.\n\n"
                        f"Formato JSON requerido:\n"
                        f'{{"nombre_contrato":"","partes":{{"empresa_a":"","empresa_b":""}},'
                        f'"fecha_inicio":"","fecha_vencimiento":"","duracion":"","renovacion_automatica":false,'
                        f'"dias_aviso_no_renovacion":"","monto_contratado":"","moneda":"",'
                        f'"forma_de_pago":"","penalidades":[],"obligaciones_empresa_a":[],'
                        f'"obligaciones_empresa_b":[],"clausulas_destacadas":[],'
                        f'"nivel_riesgo":"Bajo|Medio|Alto","justificacion_riesgo":"",'
                        f'"alertas":[],"resumen_ejecutivo":""}}\n\n'
                        f"Instrucciones:\n"
                        f"- penalidades: lista de strings describiendo cada penalidad o multa\n"
                        f"- obligaciones_empresa_a y _b: lista de strings, máximo 5 cada una\n"
                        f"- clausulas_destacadas: lista de strings con las cláusulas más importantes\n"
                        f"- alertas: lista de strings con riesgos, vencimientos próximos o condiciones desfavorables\n"
                        f"- nivel_riesgo: 'Bajo' si el contrato es favorable, 'Medio' si tiene cláusulas a negociar, 'Alto' si tiene condiciones muy desventajosas\n"
                        f"- Si un dato no aparece: escribe 'No especifica'\n\n"
                        f"Contrato:\n{preparar_texto_documento(texto_contrato)}"
                    )
                    msg = client.messages.create(
                        model="claude-haiku-4-5-20251001", max_tokens=2000,
                        messages=[{"role": "user", "content": prompt_contrato}]
                    )
                    resultado = limpiar_json(msg.content[0].text)
                    resultado["_archivo"] = archivo.name
                    st.session_state.contrato_resultado = resultado
                    st.success("Análisis completado.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al analizar el contrato: {e}")

    with tab_resultado:
        res = st.session_state.contrato_resultado
        if res is None:
            st.info("Sube y analiza un contrato para ver los resultados aquí.")
        else:
            nivel = res.get("nivel_riesgo", "Medio")
            color_riesgo = {"Bajo": "#3D7A4D", "Medio": "#B47C2B", "Alto": "#B03A2E"}.get(nivel, "#B47C2B")
            bg_riesgo    = {"Bajo": "#EAF3DE", "Medio": "#FAEEDA", "Alto": "#FCEBEB"}.get(nivel, "#FAEEDA")

            st.markdown(f"""
            <div style="background:#fff;border:1px solid #E5DFD3;border-radius:12px;padding:16px 20px;margin-bottom:14px;">
              <div style="display:flex;justify-content:space-between;align-items:flex-start;">
                <div>
                  <div style="font-weight:600;font-size:15px;color:#2E3A45;">{esc(res.get('nombre_contrato','Contrato'))}</div>
                  <div style="font-size:12px;color:#9CA8B0;margin-top:2px;">📎 {esc(res.get('_archivo',''))}</div>
                </div>
                <span style="background:{bg_riesgo};color:{color_riesgo};padding:4px 12px;border-radius:20px;font-size:12px;font-weight:600;">
                  Riesgo {esc(nivel)}
                </span>
              </div>
              <p style="font-size:13px;color:#2E3A45;margin-top:10px;line-height:1.6;">{esc(res.get('resumen_ejecutivo',''))}</p>
            </div>""", unsafe_allow_html=True)

            col1, col2 = st.columns(2)
            with col1:
                st.markdown("**Datos clave**")
                datos_clave = [
                    ("Partes", f"{esc(res.get('partes',{}).get('empresa_a',''))} ↔ {esc(res.get('partes',{}).get('empresa_b',''))}"),
                    ("Inicio", esc(res.get("fecha_inicio", "No especifica"))),
                    ("Vencimiento", esc(res.get("fecha_vencimiento", "No especifica"))),
                    ("Duración", esc(res.get("duracion", "No especifica"))),
                    ("Renovación automática", "Sí" if res.get("renovacion_automatica") else "No"),
                    ("Aviso no renovación", esc(res.get("dias_aviso_no_renovacion", "No especifica"))),
                    ("Monto", esc(res.get("monto_contratado", "No especifica"))),
                    ("Moneda", esc(res.get("moneda", "No especifica"))),
                    ("Forma de pago", esc(res.get("forma_de_pago", "No especifica"))),
                ]
                for lbl, val in datos_clave:
                    st.markdown(f"""<div style="display:flex;justify-content:space-between;padding:5px 0;
                    border-bottom:1px solid #F0EBE3;font-size:13px;">
                    <span style="color:#7A8590;">{lbl}</span>
                    <span style="font-weight:500;color:#2E3A45;text-align:right;max-width:55%;">{val}</span>
                    </div>""", unsafe_allow_html=True)
            with col2:
                st.markdown("**Alertas**")
                alertas = res.get("alertas", [])
                penalidades = res.get("penalidades", [])
                if alertas:
                    for a in alertas:
                        st.markdown(f"""<div style="background:#FAEEDA;border-left:3px solid #EF9F27;
                        border-radius:0 6px 6px 0;padding:8px 12px;margin-bottom:6px;font-size:12px;color:#854F0B;">
                        ⚠️ {esc(a)}</div>""", unsafe_allow_html=True)
                else:
                    st.markdown('<div style="color:#9CA8B0;font-size:13px;">Sin alertas detectadas</div>', unsafe_allow_html=True)
                if penalidades:
                    st.markdown("**Penalidades**")
                    for p in penalidades:
                        st.markdown(f"""<div style="background:#FCEBEB;border-left:3px solid #E24B4A;
                        border-radius:0 6px 6px 0;padding:8px 12px;margin-bottom:6px;font-size:12px;color:#A32D2D;">
                        🔴 {esc(p)}</div>""", unsafe_allow_html=True)

            st.markdown("---")
            col3, col4 = st.columns(2)
            with col3:
                obs_a = res.get("obligaciones_empresa_a", [])
                partes_str = res.get("partes", {})
                st.markdown(f"**Obligaciones — {esc(partes_str.get('empresa_a','Empresa A'))}**")
                for o in obs_a:
                    st.markdown(f"• {esc(o)}")
            with col4:
                obs_b = res.get("obligaciones_empresa_b", [])
                st.markdown(f"**Obligaciones — {esc(partes_str.get('empresa_b','Empresa B'))}**")
                for o in obs_b:
                    st.markdown(f"• {esc(o)}")

            clausulas = res.get("clausulas_destacadas", [])
            if clausulas:
                st.markdown("---")
                st.markdown("**Cláusulas destacadas**")
                for c in clausulas:
                    st.markdown(f"""<div style="background:#fff;border:1px solid #E5DFD3;border-left:3px solid #4A90B8;
                    border-radius:0 8px 8px 0;padding:8px 14px;margin-bottom:6px;font-size:13px;color:#2E3A45;">
                    {esc(c)}</div>""", unsafe_allow_html=True)

            if res.get("justificacion_riesgo"):
                st.markdown("---")
                st.info(f"**Justificación del nivel de riesgo:** {esc(res.get('justificacion_riesgo',''))}")

            resumen_export = f"""ANÁLISIS DE CONTRATO — NEXORA
{'='*50}
Contrato: {res.get('nombre_contrato','')}
Archivo:  {res.get('_archivo','')}
Nivel de riesgo: {nivel}

DATOS CLAVE
{'─'*30}
Partes:       {res.get('partes',{}).get('empresa_a','')} ↔ {res.get('partes',{}).get('empresa_b','')}
Inicio:       {res.get('fecha_inicio','')}
Vencimiento:  {res.get('fecha_vencimiento','')}
Duración:     {res.get('duracion','')}
Monto:        {res.get('monto_contratado','')} {res.get('moneda','')}
Forma de pago:{res.get('forma_de_pago','')}

RESUMEN EJECUTIVO
{'─'*30}
{res.get('resumen_ejecutivo','')}

ALERTAS
{'─'*30}
{chr(10).join('• ' + a for a in res.get('alertas', []))}

PENALIDADES
{'─'*30}
{chr(10).join('• ' + p for p in res.get('penalidades', []))}

OBLIGACIONES {res.get('partes',{}).get('empresa_a','')}
{'─'*30}
{chr(10).join('• ' + o for o in res.get('obligaciones_empresa_a', []))}

OBLIGACIONES {res.get('partes',{}).get('empresa_b','')}
{'─'*30}
{chr(10).join('• ' + o for o in res.get('obligaciones_empresa_b', []))}

JUSTIFICACIÓN DEL RIESGO
{'─'*30}
{res.get('justificacion_riesgo','')}
"""
            st.download_button("📥 Descargar resumen (.txt)", data=resumen_export.encode("utf-8"),
                               file_name=f"Nexora_Contrato_{res.get('_archivo','').replace('.pdf','').replace('.docx','')}.txt",
                               mime="text/plain", use_container_width=True)

# ── MODULO 4: GENERADOR DE RFQ ───────────────────────────────────────────────
elif st.session_state.modulo_activo == "rfq":
    titulo_modulo("Generador de RFQ", "Solicitudes de cotización profesionales en segundos")
    client = anthropic.Anthropic(api_key=st.secrets["CLAUDE_API_KEY"])

    if "rfq_generado" not in st.session_state:
        st.session_state.rfq_generado = None

    tab_form, tab_rfq = st.tabs(["📋 Completar formulario", "📄 RFQ generado"])

    with tab_form:
        c1, c2 = st.columns(2)
        with c1:
            rfq_empresa       = st.text_input("Nombre de tu empresa", placeholder="Ej: Corporación ABC SAC")
            rfq_producto      = st.text_input("Producto o servicio requerido", placeholder="Ej: Servicio de limpieza industrial")
            rfq_cantidad      = st.text_input("Cantidad / alcance", placeholder="Ej: 50 unidades / 3 sedes")
            rfq_plazo         = st.text_input("Plazo de entrega requerido", placeholder="Ej: 30 días hábiles")
            rfq_lugar         = st.text_input("Lugar de entrega / sede", placeholder="Ej: Lima, Av. Javier Prado 1234")
        with c2:
            rfq_presupuesto   = st.selectbox("Presupuesto referencial",
                ["No especificado", "< $5,000", "$5,000 - $20,000", "$20,000 - $100,000",
                 "$100,000 - $500,000", "> $500,000"])
            rfq_moneda        = st.selectbox("Moneda preferida", ["PEN (Soles)", "USD (Dólares)", "EUR (Euros)"])
            rfq_condicion_pago = st.selectbox("Condición de pago preferida",
                ["30 días desde factura", "50% adelanto / 50% entrega", "Contra entrega",
                 "Crédito 60 días", "A convenir"])
            rfq_idioma        = st.selectbox("Idioma del documento", ["Español", "Inglés", "Portugués"])
            rfq_fecha_limite  = st.date_input("Fecha límite de respuesta")
        rfq_requisitos    = st.text_area("Certificaciones o requisitos técnicos",
                            placeholder="Ej: ISO 9001, técnicos certificados, garantía mínima 12 meses")
        rfq_condiciones   = st.text_area("Condiciones especiales",
                            placeholder="Ej: Trabajo en horario nocturno, proveedor debe traer herramientas propias")
        rfq_contacto      = st.text_input("Email de contacto para respuestas", placeholder="compras@empresa.com")

        if st.button("Generar RFQ con Claude", type="primary", use_container_width=True):
            if not rfq_producto.strip():
                st.error("Completa al menos el campo 'Producto o servicio requerido'.")
            else:
                with st.spinner("Generando documento RFQ..."):
                    try:
                        prompt_rfq = (
                            f"Eres un experto en compras corporativas. Genera una Solicitud de Cotización (RFQ) "
                            f"profesional y completa en {rfq_idioma}.\n\n"
                            f"Datos del requerimiento:\n"
                            f"- Empresa solicitante: {rfq_empresa or 'No especificada'}\n"
                            f"- Producto/servicio: {rfq_producto}\n"
                            f"- Cantidad/alcance: {rfq_cantidad or 'No especificado'}\n"
                            f"- Lugar de entrega: {rfq_lugar or 'No especificado'}\n"
                            f"- Plazo de entrega requerido: {rfq_plazo or 'No especificado'}\n"
                            f"- Presupuesto referencial: {rfq_presupuesto}\n"
                            f"- Moneda: {rfq_moneda}\n"
                            f"- Condición de pago preferida: {rfq_condicion_pago}\n"
                            f"- Fecha límite de respuesta: {rfq_fecha_limite.strftime('%d de %B de %Y')}\n"
                            f"- Requisitos técnicos/certificaciones: {rfq_requisitos or 'No especificados'}\n"
                            f"- Condiciones especiales: {rfq_condiciones or 'Ninguna'}\n"
                            f"- Email de contacto: {rfq_contacto or 'No especificado'}\n\n"
                            f"Genera el documento RFQ completo con: encabezado formal, descripción del requerimiento, "
                            f"especificaciones técnicas, requisitos del proveedor, formato requerido de respuesta, "
                            f"criterios de evaluación, condiciones de pago, fecha límite y datos de contacto. "
                            f"El documento debe ser profesional, claro y listo para enviar a proveedores. "
                            f"Devuelve SOLO el texto del documento, sin comentarios adicionales."
                        )
                        msg = client.messages.create(
                            model="claude-haiku-4-5-20251001", max_tokens=2000,
                            messages=[{"role": "user", "content": prompt_rfq}]
                        )
                        st.session_state.rfq_generado = {
                            "texto": msg.content[0].text,
                            "empresa": rfq_empresa,
                            "producto": rfq_producto,
                            "fecha": rfq_fecha_limite.strftime("%Y%m%d"),
                        }
                        st.success("RFQ generado correctamente.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error al generar el RFQ: {e}")

    with tab_rfq:
        rfq = st.session_state.rfq_generado
        if rfq is None:
            st.info("Completa el formulario y haz click en 'Generar RFQ' para ver el documento aquí.")
        else:
            st.markdown(f"""
            <div style="background:#fff;border:1px solid #E5DFD3;border-left:4px solid #4A90B8;
            border-radius:0 12px 12px 0;padding:16px 20px;margin-bottom:16px;">
              <div style="font-weight:600;color:#2E3A45;font-size:14px;">
                RFQ — {esc(rfq.get('producto',''))}
              </div>
              <div style="font-size:12px;color:#9CA8B0;margin-top:2px;">
                {esc(rfq.get('empresa',''))} · Generado con Nexora IA
              </div>
            </div>""", unsafe_allow_html=True)

            st.markdown(f"""<div style="background:#F9F7F3;border:1px solid #E5DFD3;border-radius:12px;
            padding:20px 24px;font-size:13px;line-height:1.85;color:#2E3A45;
            font-family:'IBM Plex Mono',monospace;white-space:pre-wrap;">{esc(rfq.get('texto',''))}</div>""",
            unsafe_allow_html=True)

            col_a, col_b = st.columns(2)
            with col_a:
                st.download_button(
                    "📥 Descargar RFQ (.txt)",
                    data=rfq.get("texto", "").encode("utf-8"),
                    file_name=f"RFQ_{rfq.get('empresa','Empresa').replace(' ','_')}_{rfq.get('fecha','')}.txt",
                    mime="text/plain", use_container_width=True
                )
            with col_b:
                if st.button("🔄 Generar nuevo RFQ", use_container_width=True):
                    st.session_state.rfq_generado = None
                    st.rerun()

# ── MODULO 5: ANÁLISIS DE FACTURAS ──────────────────────────────────────────
elif st.session_state.modulo_activo == "facturas":
    titulo_modulo("Análisis de Facturas", "Extrae datos de facturas y exporta a Excel")
    client = anthropic.Anthropic(api_key=st.secrets["CLAUDE_API_KEY"])

    if "facturas_df" not in st.session_state:
        st.session_state.facturas_df = None

    tab_fup, tab_ftab = st.tabs(["📤 Cargar facturas", "📊 Tabla de resultados"])

    with tab_fup:
        archivos_fac = st.file_uploader(
            "Sube tus facturas (PDF)", type=["pdf"],
            accept_multiple_files=True,
            help="Puedes subir varias facturas a la vez"
        )
        st.info("Claude extrae automáticamente: proveedor, N° factura, fecha, montos sin y con IGV, ítems y condiciones de pago.")

        if archivos_fac and st.button("Analizar facturas", type="primary", use_container_width=True):
            resultados_fac = []
            pb_fac = st.progress(0)
            stxt_fac = st.empty()
            for idx, fac in enumerate(archivos_fac):
                stxt_fac.markdown(f"Analizando **{fac.name}** ({idx+1}/{len(archivos_fac)})...")
                try:
                    reader_fac = PdfReader(fac)
                    texto_fac  = "\n".join(p.extract_text() or "" for p in reader_fac.pages)

                    prompt_fac = (
                        f"Eres un experto en contabilidad. Analiza esta factura y responde EXCLUSIVAMENTE con JSON válido.\n\n"
                        f"Formato JSON requerido:\n"
                        f'{{"numero_factura":"","fecha_emision":"","fecha_vencimiento":"",'
                        f'"proveedor":"","ruc_proveedor":"","cliente":"","ruc_cliente":"",'
                        f'"descripcion_principal":"","cantidad":1,"unidad":"",'
                        f'"precio_unitario_sin_igv":0,"subtotal_sin_igv":0,'
                        f'"igv":0,"total_con_igv":0,"moneda":"",'
                        f'"condicion_pago":"","orden_compra":"","observaciones":""}}\n\n'
                        f"Instrucciones:\n"
                        f"- Todos los montos deben ser números (sin símbolos ni comas de miles, usa punto decimal)\n"
                        f"- Si un dato no aparece: escribe 'No especifica' (o 0 para montos)\n"
                        f"- descripcion_principal: el ítem o servicio principal de la factura\n\n"
                        f"Factura:\n{preparar_texto_documento(texto_fac)}"
                    )
                    msg_fac = client.messages.create(
                        model="claude-haiku-4-5-20251001", max_tokens=1000,
                        messages=[{"role": "user", "content": prompt_fac}]
                    )
                    datos_fac = limpiar_json(msg_fac.content[0].text)
                    datos_fac["Archivo"] = fac.name
                    resultados_fac.append(datos_fac)
                except Exception as e:
                    resultados_fac.append({"Archivo": fac.name, "Error": str(e)})
                pb_fac.progress((idx + 1) / len(archivos_fac))

            stxt_fac.empty()
            pb_fac.empty()
            st.session_state.facturas_df = pd.DataFrame(resultados_fac)
            st.success(f"Se analizaron {len(resultados_fac)} factura(s).")
            st.rerun()

    with tab_ftab:
        df_fac = st.session_state.facturas_df
        if df_fac is None or df_fac.empty:
            st.info("Sube y analiza facturas para ver los resultados aquí.")
        else:
            col_met1, col_met2, col_met3, col_met4 = st.columns(4)
            total_sin = pd.to_numeric(df_fac.get("subtotal_sin_igv", pd.Series([])), errors="coerce").sum()
            total_igv = pd.to_numeric(df_fac.get("igv", pd.Series([])), errors="coerce").sum()
            total_con = pd.to_numeric(df_fac.get("total_con_igv", pd.Series([])), errors="coerce").sum()
            with col_met1: st.metric("Facturas analizadas", len(df_fac))
            with col_met2: st.metric("Total sin IGV", f"{total_sin:,.2f}")
            with col_met3: st.metric("IGV total", f"{total_igv:,.2f}")
            with col_met4: st.metric("Total con IGV", f"{total_con:,.2f}")

            columnas_mostrar = [c for c in [
                "Archivo", "numero_factura", "fecha_emision", "proveedor",
                "descripcion_principal", "subtotal_sin_igv", "igv", "total_con_igv",
                "moneda", "condicion_pago"
            ] if c in df_fac.columns]
            st.dataframe(df_fac[columnas_mostrar], use_container_width=True, hide_index=True)

            def exportar_excel_facturas(df):
                buf = io.BytesIO()
                with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False, sheet_name="Facturas")
                    ws = writer.sheets["Facturas"]
                    from openpyxl.styles import Font, PatternFill, Alignment
                    header_font = Font(bold=True, color="FFFFFF", size=11)
                    header_fill = PatternFill("solid", fgColor="2E3A45")
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    for col in ws.columns:
                        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
                return buf.getvalue()

            col_dl1, col_dl2 = st.columns(2)
            with col_dl1:
                st.download_button(
                    "📥 Descargar Excel",
                    data=exportar_excel_facturas(df_fac),
                    file_name="Nexora_Facturas.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            with col_dl2:
                if st.button("🗑️ Limpiar resultados", use_container_width=True):
                    st.session_state.facturas_df = None
                    st.rerun()
